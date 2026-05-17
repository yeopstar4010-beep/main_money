"""
patch_v3b.py  —  자산계획_v3.xlsx 전체 수정 패치
실제 레이아웃: B=민엽이름, C=민엽값, D=현지이름, E=현지값, F=합계

수정 항목:
  1. 대시보드 H10 날짜: 2027-08-01 → 2027-07-01
  2. 대시보드 R11 순자산: 가용자산(C38) 참조 → 순자산(C37) 참조 [버그 수정]
  3. 대시보드 R28 여유/부족: F27-F26 → F27-F22 [버그 수정]
  4. 월별저축 B9, B10 이벤트 날짜: 2026-09-01 → 2027-09-01
  5. 자산현황 주택청약: C7 값(16,320,000) → C23(IRP·연금저축)으로 이동
  6. 자산현황 부채 R32: '대출' → '대출(어머니)'
  7. 자산현황 부채 3개 행 추가 (R33에 삽입): 주담대·보증금반환·신용
     - 소계·순자산·가용자산 수식 갱신
     - 대시보드 부채·순자산·가용자산 참조 행 번호 갱신
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime, sys
sys.stdout.reconfigure(encoding="utf-8")

PATH = r"G:\내 드라이브\claude\자산관리\자산계획_v3.xlsx"
wb = openpyxl.load_workbook(PATH)

C_NAVY, C_DGRAY = "1F3864", "595959"
C_YELLOW, C_WHITE, C_GRAY = "FFF2CC", "FFFFFF", "F2F2F2"
NUM = "#,##0"

def fl(c): return PatternFill("solid", fgColor=c)
def fn(c=C_NAVY, b=False, s=10): return Font(name="맑은 고딕", color=c, bold=b, size=s)
def bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
AL_L = Alignment(horizontal="left",   vertical="center")
AL_R = Alignment(horizontal="right",  vertical="center")
BORDER = bdr()

def set_lbl(ws, r, c, v):
    cc = ws.cell(r, c, v)
    cc.fill = fl(C_GRAY); cc.font = fn(C_DGRAY)
    cc.border = BORDER; cc.alignment = AL_L

def set_inp(ws, r, c, v=None):
    cc = ws.cell(r, c, v)
    cc.fill = fl(C_WHITE); cc.font = fn()
    cc.border = BORDER; cc.alignment = AL_R
    cc.number_format = NUM

def set_auto(ws, r, c, formula):
    cc = ws.cell(r, c, formula)
    cc.fill = fl(C_WHITE); cc.font = fn()
    cc.border = BORDER; cc.alignment = AL_R
    cc.number_format = NUM

# ── 1. 대시보드 날짜 수정 ──────────────────────────────────
ws_dash = wb["대시보드"]
c_date = ws_dash.cell(10, 8)  # H10
if hasattr(c_date.value, 'month') and c_date.value.month == 8:
    c_date.value = datetime.datetime(2027, 7, 1)
    print(f"  [1] 대시보드 H10 날짜 수정 → 2027-07-01")
else:
    print(f"  [1] 대시보드 H10 이미 수정됨: {c_date.value}")

# ── 2. 대시보드 순자산 버그 수정 (C38→C37 등) ─────────────
# 현재: C11='=자산현황!C38', E11='=자산현황!E38', F11='=자산현황!F38'
# 수정: C37=순자산, E37=순자산, F37=순자산
# (부채 3행 삽입 후 R37→R40이 되므로 나중에 또 업데이트 필요)
# 일단 R37로 수정 후, 삽입 후 R40으로 재수정
ws_dash.cell(11, 3).value = "='자산현황'!C37"
ws_dash.cell(11, 4).value = "='자산현황'!E37"
ws_dash.cell(11, 5).value = "='자산현황'!F37"
print("  [2] 대시보드 R11 순자산 참조 수정 → C/E/F 37")

# ── 3. 대시보드 여유/부족 버그 수정 ─────────────────────────
# R28 = F27-F26 (잘못됨) → F27-F22 (가용합계-지출합계)
c_surplus = ws_dash.cell(28, 6)
c_surplus.value = "=F27-F22"
print("  [3] 대시보드 R28 여유/부족 수식 수정 → F27-F22")

# ── 4. 월별저축 이벤트 날짜 수정 ──────────────────────────
ws_monthly = wb["월별저축"]
for r in [9, 10]:
    bc = ws_monthly.cell(r, 2)
    if hasattr(bc.value, 'year') and bc.value.year == 2026 and bc.value.month == 9:
        name = ws_monthly.cell(r, 3).value or ""
        bc.value = datetime.datetime(2027, 9, 1)
        print(f"  [4] 월별저축 R{r} 날짜 수정 → 2027-09-01 ({name})")

# ── 5. 자산현황 주택청약 이동 ───────────────────────────────
ws_assets = wb["자산현황"]
# C7=16320000 (민엽 주택청약) → C23(IRP·연금저축) 로 이동
c7 = ws_assets.cell(7, 3)
c23 = ws_assets.cell(23, 3)
if isinstance(c7.value, (int, float)) and c7.value > 0:
    old_c23 = c23.value if isinstance(c23.value, (int, float)) else 0
    new_c23 = old_c23 + c7.value
    print(f"  [5] 주택청약 민엽 {c7.value:,.0f} → C23 합산 ({old_c23:,.0f} + {c7.value:,.0f} = {new_c23:,.0f})")
    c23.value = new_c23
    c7.value = 0
    # B7 레이블: 주택청약 → 주택청약(연금이동)
    ws_assets.cell(7, 2).value = "주택청약(→연금)"
    # B23 레이블 업데이트
    ws_assets.cell(23, 2).value = "IRP·연금·청약"
else:
    print(f"  [5] C7 청약 값 없음 (이미 0 또는 수식): {c7.value}")

# 현지 E7=0 이지만 소계(E16)에서 이미 제외됨 → 레이블만 정리
ws_assets.cell(7, 4).value = "주택청약(→연금)"
ws_assets.cell(23, 4).value = "연금·청약(현지)"
print("  [5] 자산현황 청약 레이블 업데이트 완료")

# ── 6. 부채 이름 수정 ─────────────────────────────────────
ws_assets.cell(32, 2).value = "대출(어머니)"
print("  [6] 자산현황 R32 부채명 수정: 대출 → 대출(어머니)")

# ── 7. 부채 3개 행 삽입 (R33 앞) ─────────────────────────
# 삽입 전: R32=대출(어머니), R33=소계, R35(빈칸), R37=순자산합계섹션, R37=순자산, R38=가용자산
# 삽입 후: R33~35=신규대출, R36=소계, R40=순자산, R41=가용자산
ws_assets.insert_rows(33, 3)
print("  [7] 자산현황 R33에 3행 삽입")

new_debts = [
    ("대출(주담대)",        "대출(주담대)",        0,  0),
    ("대출(보증금 반환)",   "대출(보증금 반환)",   0,  0),
    ("대출(신용)",          "대출(신용)",          0,  0),
]
for i, (mn, dn, mv, ev) in enumerate(new_debts):
    r_new = 33 + i
    set_lbl(ws_assets, r_new, 2, mn)      # B: 민엽 이름
    ws_assets.cell(r_new, 3).border = BORDER
    set_inp(ws_assets, r_new, 3, mv)       # C: 민엽 값
    set_lbl(ws_assets, r_new, 4, dn)      # D: 현지 이름
    set_inp(ws_assets, r_new, 5, ev)       # E: 현지 값
    set_auto(ws_assets, r_new, 6, f"=C{r_new}+E{r_new}")  # F: 합계
    print(f"  [7] R{r_new}: {mn}")

# 소계 수식 갱신 (new R36, 포함 범위: C32:C35)
new_subtotal = 36
ws_assets.cell(new_subtotal, 3).value = "=SUM(C32:C35)"
ws_assets.cell(new_subtotal, 5).value = "=SUM(E32:E35)"
ws_assets.cell(new_subtotal, 6).value = f"=C{new_subtotal}+E{new_subtotal}"
print(f"  [7] 소계 수식 갱신 → R{new_subtotal}")

# 순자산 수식 갱신 (new R40: =C16+C25+C29+C36)
# (insert_rows가 같은 시트 수식은 자동 이동하므로 참조 행도 이동됨)
# 하지만 삽입 후 실제 행 확인
net_row = 40
avail_row = 41
# 순자산 수식 명시적으로 재작성 (안전하게)
ws_assets.cell(net_row, 3).value = "=C16+C25+C29+C36"
ws_assets.cell(net_row, 4).value = "=D16+D25+D29+D36"
ws_assets.cell(net_row, 5).value = "=E16+E25+E29+E36"
ws_assets.cell(net_row, 6).value = "=F16+F25+F29+F36"
print(f"  [7] 순자산 수식 갱신 → R{net_row}")

# 가용자산 수식 갱신 (new R41: =C16+C25+C36)
ws_assets.cell(avail_row, 3).value = "=C16+C25+C36"
ws_assets.cell(avail_row, 4).value = "=D16+D25+D36"
ws_assets.cell(avail_row, 5).value = "=E16+E25+E36"
ws_assets.cell(avail_row, 6).value = f"=C{avail_row}+E{avail_row}"
print(f"  [7] 가용자산 수식 갱신 → R{avail_row}")

# ── 8. 대시보드 참조 행 번호 갱신 (삽입으로 +3) ─────────────
# 부채 소계: 이전 R33 → 새 R36
ws_dash.cell(10, 3).value = "='자산현황'!C36"
ws_dash.cell(10, 4).value = "='자산현황'!E36"
ws_dash.cell(10, 5).value = "='자산현황'!F36"
print("  [8] 대시보드 R10 부채 참조 수정 → C/E/F 36")

# 순자산: 이제 R40 (버그 수정 + 행 이동)
ws_dash.cell(11, 3).value = "='자산현황'!C40"
ws_dash.cell(11, 4).value = "='자산현황'!E40"
ws_dash.cell(11, 5).value = "='자산현황'!F40"
print("  [8] 대시보드 R11 순자산 참조 수정 → C/E/F 40")

# 가용자산: 이전 F38 → 새 F41
ws_dash.cell(23, 6).value = "='자산현황'!F41"
print("  [8] 대시보드 R23 가용자산 참조 수정 → F41")

wb.save(PATH)
print(f"\n패치 완료! 파일 저장됨: {PATH}")
print("\n수동 확인 사항:")
print("  - 민엽_가계부에서 '어머니 50만' 항목 직접 삭제")
print("  - 자산현황 열 너비 확인 (삽입 후 레이아웃)")
print("  - 대시보드 숫자 검증: 순자산 = 현금+투자+계약금-부채")
