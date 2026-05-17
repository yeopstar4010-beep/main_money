"""
민엽 & 현지 자산관리 v3
- 대시보드 / 자산현황 / 자산_이력 / 민엽_가계부 / 현지_가계부 / 월별저축 / 부동산계획 / 결혼계획
실행: python build_excel_v3.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

OUT_PATH = r"G:\내 드라이브\claude\자산관리\자산계획_v3.xlsx"

# ── 팔레트 ────────────────────────────────────────────────
C_NAVY, C_BLUE, C_LBLUE = "1F3864", "2E75B6", "BDD7EE"
C_YELLOW, C_GREEN, C_RED = "FFF2CC", "E2EFDA", "FFE7E7"
C_GRAY, C_WHITE, C_DGRAY = "F2F2F2", "FFFFFF", "595959"

def fl(c): return PatternFill("solid", fgColor=c)
def fn(c=C_NAVY, b=False, s=10, i=False, bold=None):
    if bold is not None: b = bold
    return Font(name="맑은 고딕", color=c, bold=b, size=s, italic=i)
def bd(color="BFBFBF", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left",   vertical="center")
AL_R = Alignment(horizontal="right",  vertical="center")
NUM, PCT, DT, YM = "#,##0", "0.0%", "YYYY-MM-DD", "YYYY-MM"

BORDER = bd()

# 월 목록 (2026-05 ~ 2028-12)
def gen_months():
    months, d = [], datetime.date(2026, 5, 1)
    while d <= datetime.date(2028, 12, 1):
        months.append(d)
        d = datetime.date(d.year + (d.month == 12), (d.month % 12) + 1, 1)
    return months

MONTHS = gen_months()

# 수입/지출 기본값 (preset)
INCOME_PRESET = {
    "2026-05":(6000000,3500000), "2026-06":(6000000,3500000),
    "2026-07":(6000000,3500000), "2026-08":(6000000,5500000),
    "2026-09":(6000000,3500000), "2026-10":(6000000,3500000),
    "2026-11":(6000000,3500000), "2026-12":(6000000,3500000),
    "2027-01":(16000000,3500000),"2027-02":(6000000,3500000),
    "2027-03":(6000000,3500000), "2027-04":(6000000,3500000),
    "2027-05":(6000000,3500000), "2027-06":(6000000,3500000),
    "2027-07":(6000000,3500000), "2027-08":(6000000,5500000),
    "2027-09":(6000000,3500000),
}

# ── 공통 헬퍼 ──────────────────────────────────────────────
def cell(ws, r, c, v=None, bg=None, fg=C_NAVY, bold=False, sz=10,
         al=None, fmt=None, bdr=BORDER, cols=1, rows=1, italic=False):
    cc = ws.cell(row=r, column=c, value=v)
    if bg: cc.fill = fl(bg)
    cc.font = fn(fg, bold, sz, italic)
    cc.border = bdr
    if al:  cc.alignment = al
    if fmt: cc.number_format = fmt
    if cols > 1 or rows > 1:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r+rows-1, end_column=c+cols-1)
    return cc

def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def rh(ws, r, h):   ws.row_dimensions[r].height = h

def ttl(ws, r, c, txt, ncols=7):
    cell(ws, r, c, txt, C_NAVY, C_WHITE, True, 13, AL_L, cols=ncols, bdr=Border())
    rh(ws, r, 36)

def sub(ws, r, c, txt, ncols=7):
    cell(ws, r, c, txt, C_BLUE, C_WHITE, True, 10, AL_L, cols=ncols, bdr=Border())
    rh(ws, r, 22)

def hdr(ws, r, c, txt, ncols=1, nrows=1):
    return cell(ws, r, c, txt, C_LBLUE, C_NAVY, True, 10, AL_C, cols=ncols, rows=nrows)

def lbl(ws, r, c, txt="", ncols=1):
    return cell(ws, r, c, txt, C_GRAY, C_DGRAY, al=AL_L, cols=ncols)

def inp(ws, r, c, v=None, fmt=NUM, ncols=1, bg=C_WHITE):
    return cell(ws, r, c, v, bg, C_NAVY, al=AL_R, fmt=fmt, cols=ncols)

def inpl(ws, r, c, v=None, ncols=1):
    return cell(ws, r, c, v, C_WHITE, C_NAVY, al=AL_L, cols=ncols)

def tot(ws, r, c, v=None, fmt=NUM, ncols=1):
    cc = cell(ws, r, c, v, C_YELLOW, C_NAVY, True, 10, AL_R, fmt, cols=ncols)
    rh(ws, r, 20)
    return cc

def note(ws, r, c, txt):
    cc = ws.cell(r, c, txt)
    cc.font = Font(name="맑은 고딕", size=9, italic=True, color="808080")
    cc.alignment = AL_L

def auto(ws, r, c, formula, fmt=NUM):
    cc = ws.cell(r, c, formula)
    cc.fill = fl(C_WHITE); cc.font = fn(); cc.border = BORDER
    cc.alignment = AL_R; cc.number_format = fmt
    return cc


# ═══════════════════════════════════════════════════════════
# 가계부 (민엽 / 현지) - 공통 빌더
# ═══════════════════════════════════════════════════════════
EXPENSE_CATS = [
    "식비/외식", "교통/주유", "통신/구독", "쇼핑/의류",
    "여가/취미", "의료/건강", "경조사", "고정지출", "부모님/용돈", "기타"
]

def build_household(wb, name, sheet_name, init_values=None):
    """가계부 시트 생성. init_values: {YYYY-MM: [c0..c9]} 선택적"""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    cw(ws, 2, 9)   # 월
    for i in range(len(EXPENSE_CATS)):
        cw(ws, 3 + i, 11)
    cw(ws, 3 + len(EXPENSE_CATS), 13)  # 합계
    cw(ws, 3 + len(EXPENSE_CATS) + 1, 16)  # 메모

    TOTAL_COL = 3 + len(EXPENSE_CATS)       # M (13)
    MEMO_COL  = TOTAL_COL + 1               # N (14)

    r = 2
    ttl(ws, r, 2, f"  {name} 월별 지출 가계부", ncols=MEMO_COL - 1); r += 1
    note(ws, r, 2, f"  각 카테고리별 지출을 입력하세요. 합계는 자동 계산 → 월별저축 시트에 자동 반영됩니다.")
    rh(ws, r, 16); r += 1

    # 헤더
    hdr(ws, r, 2, "월")
    for i, cat in enumerate(EXPENSE_CATS):
        hdr(ws, r, 3 + i, cat)
    hdr(ws, r, TOTAL_COL, "월 합계")
    hdr(ws, r, MEMO_COL, "메모")
    r += 1

    DATA_START = r  # 가계부 데이터 시작행 (반환용)

    for month in MONTHS:
        ym = month.strftime("%Y-%m")
        # 월 셀
        mc = ws.cell(r, 2, month)
        mc.number_format = YM; mc.font = fn(C_NAVY, bold=True)
        mc.fill = fl(C_GRAY); mc.border = BORDER; mc.alignment = AL_C

        vals = (init_values or {}).get(ym, [None]*len(EXPENSE_CATS))
        for i in range(len(EXPENSE_CATS)):
            inp(ws, r, 3 + i, vals[i] if i < len(vals) else None)

        # 합계
        cat_range = f"{get_column_letter(3)}{r}:{get_column_letter(TOTAL_COL-1)}{r}"
        auto(ws, r, TOTAL_COL, f"=SUM({cat_range})")
        # 메모
        inpl(ws, r, MEMO_COL)
        r += 1

    DATA_END = r - 1

    r += 1
    # 카테고리별 연간 합계
    sub(ws, r, 2, "카테고리별 합계", ncols=MEMO_COL - 1); r += 1
    lbl(ws, r, 2, "전체 합계")
    for i in range(len(EXPENSE_CATS)):
        col_letter = get_column_letter(3 + i)
        tot(ws, r, 3 + i, f"=SUM({col_letter}{DATA_START}:{col_letter}{DATA_END})")
    tot(ws, r, TOTAL_COL,
        f"=SUM({get_column_letter(TOTAL_COL)}{DATA_START}:{get_column_letter(TOTAL_COL)}{DATA_END})")
    r += 2

    note(ws, r, 2, "  ※ 각 셀에 숫자만 입력. 월 합계는 자동 계산.")

    return ws, DATA_START, DATA_END, TOTAL_COL


# ═══════════════════════════════════════════════════════════
# 월별저축
# ═══════════════════════════════════════════════════════════
def build_monthly(wb, m_sheet, h_sheet, m_data_start, h_data_start, m_total_col, h_total_col):
    ws = wb.create_sheet("월별저축")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    cw(ws, 2, 9); cw(ws, 3, 14); cw(ws, 4, 14); cw(ws, 5, 14)
    cw(ws, 6, 14); cw(ws, 7, 14); cw(ws, 8, 12); cw(ws, 9, 20)

    r = 2
    ttl(ws, r, 2, "  월별 저축 추이", ncols=8); r += 1
    note(ws, r, 2, "  수입만 입력하세요. 지출은 가계부 시트에서 자동 반영됩니다.")
    rh(ws, r, 16); r += 1

    # 기준 순자산
    sub(ws, r, 2, "시작 순자산 (2026년 5월)", ncols=8); r += 1
    lbl(ws, r, 2, "시작 순자산")
    BASE_ROW = r
    inp(ws, r, 3, 450986632, ncols=2)
    inpl(ws, r, 5, "  자산현황 시트 순자산 값 입력", ncols=4)
    r += 2

    # 이벤트
    sub(ws, r, 2, "주요 이벤트 (해당 월 자동 차감)", ncols=8); r += 1
    hdr(ws, r, 2, "연-월"); hdr(ws, r, 3, "이벤트명", ncols=3)
    hdr(ws, r, 6, "금액 (원)"); hdr(ws, r, 7, "비고", ncols=2); r += 1

    events = [
        ("2027-08", "어머니 차용금 상환",    -106000000),
        ("2027-08", "결혼식 비용",            -50000000),
        ("2027-08", "인테리어·가전가구",      -40000000),
        ("2027-09", "부동산 잔금 (자기자금)", -220000000),
        ("2027-09", "취득세·부대비용",        -50000000),
        ("2027-09", "신혼여행",               -10000000),
        ("2027-09", "세입자 보증금 반환",     -300000000),
    ]
    EVENT_ROWS = []
    for ev_ym, ev_name, ev_amt in events:
        inpl(ws, r, 2, ev_ym)
        inpl(ws, r, 3, ev_name, ncols=3)
        inp(ws, r, 6, ev_amt, bg=C_RED if ev_amt < 0 else C_GREEN)
        inpl(ws, r, 7, ncols=2)
        EVENT_ROWS.append(r); r += 1
    for _ in range(3):
        inpl(ws, r, 2); inpl(ws, r, 3, ncols=3); inp(ws, r, 6); inpl(ws, r, 7, ncols=2)
        EVENT_ROWS.append(r); r += 1
    EV_START, EV_END = EVENT_ROWS[0], EVENT_ROWS[-1]
    r += 1

    # 월별 테이블
    sub(ws, r, 2, "월별 기록", ncols=8); r += 1
    for c, txt in zip([2,3,4,5,6,7,8,9],
                      ["월","민엽 수입","현지 수입","합계 수입",
                       "합계 지출","저축액","저축률","비고"]):
        hdr(ws, r, c, txt)
    r += 1
    DATA_START = r

    for idx, month in enumerate(MONTHS):
        ym = month.strftime("%Y-%m")
        p = INCOME_PRESET.get(ym, (None, None))

        mc = ws.cell(r, 2, month)
        mc.number_format = YM; mc.font = fn(C_NAVY, bold=True)
        mc.fill = fl(C_GRAY); mc.border = BORDER; mc.alignment = AL_C

        inp(ws, r, 3, p[0])  # 민엽 수입
        inp(ws, r, 4, p[1])  # 현지 수입

        # 합계 수입
        auto(ws, r, 5, f"=C{r}+D{r}")

        # 합계 지출 = 민엽_가계부 + 현지_가계부 (SUMIF로 월 매칭)
        m_col = get_column_letter(m_total_col)
        h_col = get_column_letter(h_total_col)
        m_month_col = "B"  # 가계부 시트의 월 컬럼
        exp_formula = (
            f"=SUMIF('{m_sheet}'!{m_month_col}:{m_month_col},"
            f"TEXT(B{r},\"YYYY-MM\"),"
            f"'{m_sheet}'!{m_col}:{m_col})"
            f"+SUMIF('{h_sheet}'!{m_month_col}:{m_month_col},"
            f"TEXT(B{r},\"YYYY-MM\"),"
            f"'{h_sheet}'!{h_col}:{h_col})"
        )
        auto(ws, r, 6, exp_formula)

        # 저축
        auto(ws, r, 7, f"=E{r}-F{r}")

        # 저축률
        auto(ws, r, 8, f"=IF(E{r}=0,0,G{r}/E{r})", PCT)

        # 비고
        inpl(ws, r, 9)
        r += 1

    DATA_END = r - 1

    # 누계
    r += 1
    sub(ws, r, 2, "누계", ncols=8); r += 1
    lbl(ws, r, 2, "2026.05 ~ 2027.09 저축 합계")
    CUM_ROW = r
    tot(ws, r, 7,
        f"=SUMPRODUCT((B{DATA_START}:B{DATA_END}>=DATE(2026,5,1))"
        f"*(B{DATA_START}:B{DATA_END}<=DATE(2027,9,1))"
        f"*G{DATA_START}:G{DATA_END})")
    r += 1
    lbl(ws, r, 2, "월 평균 저축")
    tot(ws, r, 7, f"=AVERAGE(G{DATA_START}:G{DATA_END})")
    r += 2

    note(ws, r, 2, "  ※ 수입(C·D열)만 입력하세요. 지출은 가계부 시트 자동 합산, 저축·저축률 자동 계산.")
    note(ws, r+1, 2, "  ※ 이벤트는 위 표에서 수정하세요.")

    return ws, BASE_ROW, EV_START, EV_END, DATA_START, DATA_END, CUM_ROW


# ═══════════════════════════════════════════════════════════
# 자산_이력
# ═══════════════════════════════════════════════════════════
def build_history(wb):
    ws = wb.create_sheet("자산_이력")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    cw(ws, 2, 9)
    # 민엽: 현금(3), 주식(4), 코인(5), 기타투자(6), 순자산(7)
    # 현지: 현금(8), 주식(9), 기타투자(10), 순자산(11)
    # 합계: 순자산(12), 전월비(13), 변동률(14)
    for c in range(3, 13): cw(ws, c, 13)
    cw(ws, 13, 13); cw(ws, 14, 10)

    r = 2
    ttl(ws, r, 2, "  자산 월별 이력", ncols=13); r += 1
    note(ws, r, 2, "  매월 말일 잔액을 직접 입력하세요. 순자산·전월비·변동률은 자동 계산됩니다.")
    rh(ws, r, 16); r += 1

    # 헤더 (2행)
    cell(ws, r, 2, "월", C_LBLUE, C_NAVY, True, 10, AL_C, rows=2); r_hdr = r
    cell(ws, r, 3, "민  엽", C_NAVY, C_WHITE, True, 10, AL_C, cols=5, bdr=Border())
    cell(ws, r, 8, "현  지", C_BLUE, C_WHITE, True, 10, AL_C, cols=4, bdr=Border())
    cell(ws, r, 12, "합  계", C_LBLUE, C_NAVY, True, 10, AL_C, cols=3, bdr=Border())
    r += 1
    for c, txt in zip([3,4,5,6,7], ["현금·예금","주식·코인","IRP·연금","기타","순자산"]):
        hdr(ws, r, c, txt)
    for c, txt in zip([8,9,10,11], ["현금·예금","주식","IRP·연금","순자산"]):
        hdr(ws, r, c, txt)
    for c, txt in zip([12,13,14], ["순자산 합계","전월비","변동률"]):
        hdr(ws, r, c, txt)
    ws.merge_cells(start_row=r_hdr, start_column=2, end_row=r+1-1, end_column=2)
    r += 1
    DATA_START = r

    # 초기값 (2026-05)
    init = {
        "2026-05": {
            "m": [12513400, 195891642, 21144836, 9381394],
            "h": [28489076, 119486803, 2562952, 0],
        }
    }

    prev_net_row = None
    for idx, month in enumerate(MONTHS):
        ym = month.strftime("%Y-%m")
        iv = init.get(ym, {})
        mv = iv.get("m", [None]*4)
        hv = iv.get("h", [None]*4)

        mc = ws.cell(r, 2, month)
        mc.number_format = YM; mc.font = fn(C_NAVY, bold=True)
        mc.fill = fl(C_GRAY); mc.border = BORDER; mc.alignment = AL_C

        # 민엽 입력
        for i, v in enumerate(mv): inp(ws, r, 3+i, v)
        # 민엽 순자산 = 현금+주식+IRP+기타 - 부채 (어머니차용금 별도)
        auto(ws, r, 7, f"=C{r}+D{r}+E{r}+F{r}")

        # 현지 입력
        for i, v in enumerate(hv): inp(ws, r, 8+i, v)
        # 현지 순자산
        auto(ws, r, 11, f"=H{r}+I{r}+J{r}+K{r}")

        # 합계 순자산
        auto(ws, r, 12, f"=G{r}+K{r}")

        # 전월비
        if prev_net_row is None:
            auto(ws, r, 13, "")
        else:
            auto(ws, r, 13, f"=L{r}-L{prev_net_row}")

        # 변동률
        if prev_net_row is None:
            auto(ws, r, 14, "", PCT)
        else:
            auto(ws, r, 14, f"=IF(L{prev_net_row}=0,0,(L{r}-L{prev_net_row})/ABS(L{prev_net_row}))", PCT)
            ws.cell(r, 14).fill = fl(C_WHITE)

        prev_net_row = r; r += 1

    DATA_END = r - 1; r += 1
    note(ws, r, 2, "  ※ C~F(민엽), H~K(현지)열에 매월 말 잔액 입력. 순자산·전월비·변동률은 자동 계산.")
    note(ws, r+1, 2, "  ※ 민엽 부채(어머니 차용금 1억600만)는 순자산에서 별도 차감하세요.")

    return ws


# ═══════════════════════════════════════════════════════════
# 부동산계획
# ═══════════════════════════════════════════════════════════
def build_realestate(wb):
    ws = wb.create_sheet("부동산계획")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    cw(ws, 2, 24); cw(ws, 3, 4); cw(ws, 4, 18); cw(ws, 5, 18); cw(ws, 6, 18)

    r = 2
    ttl(ws, r, 2, "  부동산 계획", ncols=5); r += 1
    note(ws, r, 2, "  노란색 셀에 값을 입력하세요.")
    rh(ws, r, 16); r += 1

    # ── 요약 블록 (대시보드 참조용 고정 영역) ──────────────
    sub(ws, r, 2, "핵심 요약 (자동 계산)", ncols=5); r += 1
    SUMMARY_START = r
    summary_items = [
        ("매매가",           1210000000),
        ("계약금 납부",      130000000),
        ("잔금 자기자금",    220000000),   # D_{r+2}
        ("대출 예정",        560000000),   # D_{r+3}
        ("취득세 합계",      None),        # D_{r+4} = formula
        ("세입자 보증금",    300000000),   # D_{r+5}
        ("인테리어 합계",    None),        # D_{r+6} = formula
    ]
    SUMMARY_ROWS = {}
    for sname, sval in summary_items:
        lbl(ws, r, 2, sname); cell(ws, r, 3, "")
        inp(ws, r, 4, sval, ncols=2)
        SUMMARY_ROWS[sname] = r; r += 1
    r += 1

    # ① 자금 조달
    sub(ws, r, 2, "① 자금 조달 계획", ncols=5); r += 1
    hdr(ws, r, 2, "항  목"); cell(ws, r, 3, "")
    hdr(ws, r, 4, "금  액"); hdr(ws, r, 5, "비  고", ncols=2); r += 1
    fund_items = [
        ("계약금 납부 (완료)",    130000000, "2025.04 납부"),
        ("잔금 자기자금",         220000000, ""),
        ("대출 예정금액",         560000000, "LTV 확정 후 수정"),
        ("기타 (부모님 지원 등)", 0,         ""),
    ]
    fund_rows = []
    for fn_name, fv, fn_note in fund_items:
        lbl(ws, r, 2, fn_name); cell(ws, r, 3, "")
        inp(ws, r, 4, fv); inpl(ws, r, 5, fn_note, ncols=2)
        fund_rows.append(r); r += 1
    tot(ws, r, 2, "  합  계 (≒ 매매가 확인)"); cell(ws, r, 3, "")
    tot(ws, r, 4, f"=SUM({','.join(f'D{x}' for x in fund_rows)})")
    r += 2

    # ② 대출 상세
    sub(ws, r, 2, "② 대출 상세", ncols=5); r += 1
    loan_items = [
        ("대출 종류",       None, "예: 주택담보대출 / 디딤돌"),
        ("대출 금액 (원)",  None, ""),
        ("연 금리 (%)",     None, ""),
        ("대출 기간 (년)",  None, ""),
        ("거치 기간 (년)",  None, ""),
        ("월 상환액 (원)",  None, "은행 계산기 확인"),
        ("실행 예정일",     None, ""),
        ("만기일",          None, ""),
    ]
    for lname, lv, lnote in loan_items:
        lbl(ws, r, 2, lname); cell(ws, r, 3, "")
        inpl(ws, r, 4, lv); inpl(ws, r, 5, lnote, ncols=2); r += 1
    r += 1

    # ③ 취득세·부대비용
    sub(ws, r, 2, "③ 취득세 · 부대비용", ncols=5); r += 1
    hdr(ws, r, 2, "항  목"); cell(ws, r, 3, "")
    hdr(ws, r, 4, "예산"); hdr(ws, r, 5, "실제"); hdr(ws, r, 6, "비고"); r += 1
    tax_items = ["취득세", "농어촌특별세", "지방교육세",
                 "법무사 수수료", "부동산 중개수수료", "이사 비용", "기타"]
    tax_rows = []
    for tname in tax_items:
        lbl(ws, r, 2, tname); cell(ws, r, 3, "")
        inp(ws, r, 4); inp(ws, r, 5); inpl(ws, r, 6)
        tax_rows.append(r); r += 1
    tax_sum_formula = f"=SUM({','.join(f'D{x}' for x in tax_rows)})"
    lbl(ws, r, 2, "취득세 합계"); cell(ws, r, 3, "")
    TAX_TOTAL_ROW = r
    tot(ws, r, 4, tax_sum_formula); r += 2

    # ④ 세입자 보증금
    sub(ws, r, 2, "④ 세입자 보증금 반환", ncols=5); r += 1
    note(ws, r, 2, "  ※ 잔금과 별도 — 입주 시 세입자에게 반환해야 하는 금액입니다.")
    rh(ws, r, 16); r += 1
    lbl(ws, r, 2, "현 세입자 보증금"); cell(ws, r, 3, "")
    DEPOSIT_ROW = r
    inp(ws, r, 4, 300000000); inpl(ws, r, 5, "입주 시 반환", ncols=2); r += 1
    lbl(ws, r, 2, "반환 예정일"); cell(ws, r, 3, "")
    inpl(ws, r, 4, "2027-09 또는 입주 시기"); r += 2

    # ⑤ 인테리어·가전
    sub(ws, r, 2, "⑤ 인테리어 · 가전가구", ncols=5); r += 1
    hdr(ws, r, 2, "항  목"); cell(ws, r, 3, "")
    hdr(ws, r, 4, "예산"); hdr(ws, r, 5, "실제"); hdr(ws, r, 6, "비고"); r += 1
    int_items = ["철거·기본공사","도배·장판","주방공사","욕실공사",
                 "전기·조명","에어컨","냉장고","세탁기·건조기",
                 "소파·침대·가구","TV·가전","기타"]
    int_rows = []
    for iname in int_items:
        lbl(ws, r, 2, iname); cell(ws, r, 3, "")
        inp(ws, r, 4); inp(ws, r, 5); inpl(ws, r, 6)
        int_rows.append(r); r += 1
    int_sum_formula = f"=SUM({','.join(f'D{x}' for x in int_rows)})"
    lbl(ws, r, 2, "인테리어 합계"); cell(ws, r, 3, "")
    INT_TOTAL_ROW = r
    tot(ws, r, 4, int_sum_formula); r += 2

    # 요약 블록에 수식 연결
    ws.cell(SUMMARY_ROWS["취득세 합계"], 4).value = f"=D{TAX_TOTAL_ROW}"
    ws.cell(SUMMARY_ROWS["인테리어 합계"], 4).value = f"=D{INT_TOTAL_ROW}"

    note(ws, r, 2, "  ※ D열(예산) 입력 → 취득세·인테리어 합계가 요약 블록에 자동 반영됩니다.")

    return ws, SUMMARY_ROWS, DEPOSIT_ROW


# ═══════════════════════════════════════════════════════════
# 결혼계획
# ═══════════════════════════════════════════════════════════
def build_wedding(wb):
    ws = wb.create_sheet("결혼계획")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    cw(ws, 2, 22); cw(ws, 3, 4); cw(ws, 4, 16); cw(ws, 5, 16); cw(ws, 6, 14); cw(ws, 7, 16)

    r = 2
    ttl(ws, r, 2, "  결혼 계획", ncols=6); r += 1
    note(ws, r, 2, "  결혼식: 2027년 8월 | 신혼여행: 2027년 9월")
    rh(ws, r, 16); r += 1

    # ── 요약 블록 ──────────────────────────────────────────
    sub(ws, r, 2, "핵심 요약 (자동 계산)", ncols=6); r += 1
    SUMMARY_ROWS = {}
    for sname in ["결혼식 합계", "신혼여행 합계", "결혼 총비용", "예상 축의금", "결혼 순비용"]:
        lbl(ws, r, 2, sname); cell(ws, r, 3, "")
        inp(ws, r, 4, ncols=3)
        SUMMARY_ROWS[sname] = r; r += 1
    r += 1

    # ① 결혼식 세부
    sub(ws, r, 2, "① 결혼식 세부 예산", ncols=6); r += 1
    hdr(ws, r, 2, "항  목"); cell(ws, r, 3, "")
    hdr(ws, r, 4, "예산 (원)"); hdr(ws, r, 5, "실제 (원)")
    hdr(ws, r, 6, "차  액"); hdr(ws, r, 7, "비  고"); r += 1

    sections = [
        ("스드메", ["스튜디오 (사진촬영)", "드레스", "메이크업 (헤어포함)", "신랑 예복", "부케·부토니에"]),
        ("예식장", ["예식장 대관료", "뷔페·식사", "웨딩케이크", "사회자", "영상촬영", "청첩장·답례품"]),
        ("예물·예단", ["예물 (반지 등)", "예단 비용"]),
        ("기타", ["기타 ①", "기타 ②"]),
    ]
    all_budget_rows = []
    for sec, items in sections:
        sec_rows = []
        for item in items:
            lbl(ws, r, 2, f"  {item}"); cell(ws, r, 3, "")
            inp(ws, r, 4); inp(ws, r, 5)
            auto(ws, r, 6, f"=E{r}-D{r}"); inpl(ws, r, 7)
            sec_rows.append(r); all_budget_rows.append(r); r += 1
        lbl(ws, r, 2, f"  [{sec}] 소계"); cell(ws, r, 3, "")
        sec_d = f"=SUM({','.join(f'D{x}' for x in sec_rows)})"
        sec_e = f"=SUM({','.join(f'E{x}' for x in sec_rows)})"
        tot(ws, r, 4, sec_d); tot(ws, r, 5, sec_e); tot(ws, r, 6, f"=E{r}-D{r}"); r += 1

    all_d = f"=SUM({','.join(f'D{x}' for x in all_budget_rows)})"
    all_e = f"=SUM({','.join(f'E{x}' for x in all_budget_rows)})"
    r += 1
    WEDDING_TOTAL = r
    lbl(ws, r, 2, "결혼식 합계"); cell(ws, r, 3, "")
    tot(ws, r, 4, all_d); tot(ws, r, 5, all_e); tot(ws, r, 6, f"=E{r}-D{r}")
    r += 2

    # ② 신혼여행
    sub(ws, r, 2, "② 신혼여행", ncols=6); r += 1
    hdr(ws, r, 2, "항  목"); cell(ws, r, 3, "")
    hdr(ws, r, 4, "예산"); hdr(ws, r, 5, "실제")
    hdr(ws, r, 6, "차  액"); hdr(ws, r, 7, "비  고"); r += 1
    trip_rows = []
    for tname in ["항공권 (2인)", "숙박", "투어·액티비티", "식비·쇼핑", "기타"]:
        lbl(ws, r, 2, f"  {tname}"); cell(ws, r, 3, "")
        inp(ws, r, 4); inp(ws, r, 5)
        auto(ws, r, 6, f"=E{r}-D{r}"); inpl(ws, r, 7)
        trip_rows.append(r); r += 1
    TRIP_TOTAL = r
    lbl(ws, r, 2, "신혼여행 합계"); cell(ws, r, 3, "")
    tot(ws, r, 4, f"=SUM({','.join(f'D{x}' for x in trip_rows)})")
    tot(ws, r, 5, f"=SUM({','.join(f'E{x}' for x in trip_rows)})")
    tot(ws, r, 6, f"=E{r}-D{r}"); r += 2

    # ③ 축의금 시뮬레이터
    sub(ws, r, 2, "③ 축의금 시뮬레이터", ncols=6); r += 1
    hdr(ws, r, 2, "하객 구분"); cell(ws, r, 3, "")
    hdr(ws, r, 4, "민엽 인원"); hdr(ws, r, 5, "현지 인원")
    hdr(ws, r, 6, "평균 축의금"); hdr(ws, r, 7, "예상 수입"); r += 1
    gift_groups = [
        ("직장 동료", None, None, 70000),
        ("친구·지인", None, None, 100000),
        ("친척·가족", None, None, 200000),
        ("기타",       None, None, 100000),
    ]
    gift_rows = []
    for gname, gm, gh, gavg in gift_groups:
        lbl(ws, r, 2, gname); cell(ws, r, 3, "")
        inp(ws, r, 4, gm); inp(ws, r, 5, gh); inp(ws, r, 6, gavg)
        auto(ws, r, 7, f"=(D{r}+E{r})*F{r}")
        gift_rows.append(r); r += 1
    GIFT_TOTAL = r
    lbl(ws, r, 2, "축의금 합계"); cell(ws, r, 3, "")
    for c in [4, 5, 6]: ws.cell(r, c).border = BORDER
    tot(ws, r, 7, f"=SUM({','.join(f'G{x}' for x in gift_rows)})"); r += 2

    # 요약 블록 수식 연결
    ws.cell(SUMMARY_ROWS["결혼식 합계"],  4).value = f"=D{WEDDING_TOTAL}"
    ws.cell(SUMMARY_ROWS["신혼여행 합계"], 4).value = f"=D{TRIP_TOTAL}"
    ws.cell(SUMMARY_ROWS["결혼 총비용"],  4).value = \
        f"=D{WEDDING_TOTAL}+D{TRIP_TOTAL}"
    ws.cell(SUMMARY_ROWS["예상 축의금"],  4).value = f"=G{GIFT_TOTAL}"
    ws.cell(SUMMARY_ROWS["결혼 순비용"],  4).value = \
        f"=D{SUMMARY_ROWS['결혼 총비용']}-D{SUMMARY_ROWS['예상 축의금']}"
    for srow in SUMMARY_ROWS.values():
        ws.cell(srow, 4).fill = fl(C_YELLOW)
        ws.cell(srow, 4).font = fn(C_NAVY, bold=True)
        ws.cell(srow, 4).number_format = NUM
        ws.cell(srow, 4).alignment = AL_R
        ws.cell(srow, 4).border = BORDER

    note(ws, r, 2, "  ※ D열(예산), 하객 인원을 입력하세요. 합계·순비용은 자동 계산.")

    return ws, SUMMARY_ROWS, WEDDING_TOTAL, TRIP_TOTAL, GIFT_TOTAL


# ═══════════════════════════════════════════════════════════
# 자산현황
# ═══════════════════════════════════════════════════════════
def build_assets(wb):
    ws = wb.create_sheet("자산현황")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    cw(ws, 2, 22); cw(ws, 3, 4); cw(ws, 4, 18); cw(ws, 5, 18); cw(ws, 6, 18); cw(ws, 7, 18)

    r = 2
    ttl(ws, r, 2, "  자산 현황", ncols=6); r += 1
    note(ws, r, 2, "  기준: 매월 말일 업데이트 | D=민엽, E=현지, F=합계")
    rh(ws, r, 16); r += 1

    hdr(ws, r, 2, "항  목"); cell(ws, r, 3, "")
    hdr(ws, r, 4, "민  엽"); hdr(ws, r, 5, "현  지")
    hdr(ws, r, 6, "합  계"); hdr(ws, r, 7, "비  고"); r += 1

    def section(title, items_m, items_h):
        nonlocal r
        sub(ws, r, 2, title, ncols=6); r += 1
        rows = []
        for (mn, mv), (hn, hv) in zip(items_m, items_h):
            lbl(ws, r, 2, mn); cell(ws, r, 3, "")
            inp(ws, r, 4, mv); inp(ws, r, 5, hv)
            auto(ws, r, 6, f"=D{r}+E{r}")
            inpl(ws, r, 7, hn)
            rows.append(r); r += 1
        SUM_ROW = r
        lbl(ws, r, 2, "  소  계"); cell(ws, r, 3, "")
        tot(ws, r, 4, f"=SUM({','.join(f'D{x}' for x in rows)})")
        tot(ws, r, 5, f"=SUM({','.join(f'E{x}' for x in rows)})")
        tot(ws, r, 6, f"=D{r}+E{r}")
        cell(ws, r, 7, ""); r += 1
        return SUM_ROW

    cash_m = [("토스뱅크",1794954),("하나은행",72561),("기업은행",109094),
               ("카카오뱅크",150012),("우리은행",50000),("OK토스플러스",300000),
               ("지역화폐",36779),("배터리카드",400000),("기타",3600000)]
    cash_h = [("OK저축은행",""),("주택청약",""),("농협",""),
               ("적금①",""),("적금②",""),("기타예금",""),
               ("",""),("",""),("","")]
    cash_h_v = [20722500,7000000,766576,0,0,0,0,0,0]
    CASH_ROW = section("현금 · 예금",
        cash_m,
        [(cash_h[i][0], cash_h_v[i]) for i in range(len(cash_h))])

    inv_m = [("키움증권",87576544),("토스주식",99471104),("자사주",9381394),
              ("ISA (민엽)",9381394),("IRP·연금저축",21144836),("코인",31263027)]
    inv_h_n = ["한국투자증권","토스증권","대신증권","ISA (현지)","연금저축","업비트"]
    inv_h_v = [76114704,41075120,2035000,5297783,2562952,680298]
    INV_ROW = section("투자 자산 (주식 · 코인 · 연금)",
        inv_m, list(zip(inv_h_n, inv_h_v)))

    CONT_ROW = section("계약금 납부 (유동성 없음)",
        [("계약금 납부 (아파트)",130000000)],
        [("2025.04 납부 / 잔금 전 고정",0)])

    DEBT_ROW = section("부  채 (마이너스 입력)",
        [("어머니 차용금",-106000000)],
        [("계약금 조달 시 차용",0)])

    r += 1
    sub(ws, r, 2, "순자산 합계", ncols=6); r += 1
    NET_ROW = r
    lbl(ws, r, 2, "순자산 = 현금+투자+계약금 - 부채"); cell(ws, r, 3, "")
    for c, ref in [(4,"D"),(5,"E"),(6,"F")]:
        tot(ws, r, c,
            f"={ref}{CASH_ROW}+{ref}{INV_ROW}+{ref}{CONT_ROW}+{ref}{DEBT_ROW}")
        ws.cell(r, c).font = fn(C_NAVY, bold=True, s=11)
    rh(ws, r, 26); r += 1

    AVAIL_ROW = r
    lbl(ws, r, 2, "가용 자산 = 현금+투자 - 부채"); cell(ws, r, 3, "")
    for c, ref in [(4,"D"),(5,"E"),(6,"F")]:
        tot(ws, r, c,
            f"={ref}{CASH_ROW}+{ref}{INV_ROW}+{ref}{DEBT_ROW}")
    rh(ws, r, 22); r += 2
    note(ws, r, 2, "  ※ D·E열(금액)만 수정. F열과 소계·순자산은 자동 계산.")

    return ws, CASH_ROW, INV_ROW, CONT_ROW, DEBT_ROW, NET_ROW, AVAIL_ROW


# ═══════════════════════════════════════════════════════════
# 대시보드
# ═══════════════════════════════════════════════════════════
def build_dashboard(wb, asset_rows, monthly_rows, re_rows, wedding_rows):
    ws = wb.create_sheet("대시보드")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    for c, w in [(2,18),(3,16),(4,16),(5,14),(6,4),(7,20),(8,14),(9,14)]:
        cw(ws, c, w)

    _, _, _, _, _, NET_ROW, AVAIL_ROW = asset_rows
    _, _, _, _, CUM_ROW = monthly_rows[:5]  # BASE_ROW, EV_START, EV_END, DATA_START, DATA_END
    RE_SUMMARY_ROWS, _ = re_rows
    W_SUMMARY_ROWS, _, _, _ = wedding_rows

    r = 2
    ttl(ws, r, 2, "  민엽 & 현지  자산관리 대시보드", ncols=8); r += 1
    cell(ws, r, 2, f"  기준일: {datetime.date.today().strftime('%Y년 %m월 %d일')}  |  각 시트에 입력하면 자동 업데이트",
         C_LBLUE, C_NAVY, al=AL_L, cols=8, bdr=Border())
    rh(ws, r, 18); r += 2

    # ── 순자산 현황 ──────────────────────────────────────
    sub(ws, r, 2, "현재 순자산", ncols=4); r += 1
    for c, txt in zip([2,3,4,5], ["항목","민엽","현지","합계"]):
        hdr(ws, r, c, txt)
    r += 1
    for row_label, asset_row in [
        ("현금·예금",       asset_rows[0]),  # CASH_ROW
        ("투자자산",        asset_rows[1]),  # INV_ROW
        ("계약금",          asset_rows[2]),  # CONT_ROW
        ("부채",            asset_rows[3]),  # DEBT_ROW
    ]:
        lbl(ws, r, 2, row_label)
        for c, ref in [(3,"D"),(4,"E"),(5,"F")]:
            auto(ws, r, c, f"='자산현황'!{ref}{asset_row}")
        r += 1
    lbl(ws, r, 2, "순  자  산")
    for c, ref in [(3,"D"),(4,"E"),(5,"F")]:
        tot(ws, r, c, f"='자산현황'!{ref}{NET_ROW}")
        ws.cell(r, c).font = fn(C_NAVY, bold=True, s=11)
    rh(ws, r, 24); r += 2

    # ── D-Day ──────────────────────────────────────────
    sub(ws, r, 7, "D-Day", ncols=3); r += 1
    for ev, target in [("결혼식 (2027.08)","2027-08-01"), ("잔금일 (2027.09)","2027-09-01")]:
        lbl(ws, r, 7, ev)
        c_date = ws.cell(r, 8, datetime.date.fromisoformat(target))
        c_date.number_format = DT; c_date.fill = fl(C_WHITE)
        c_date.font = fn(); c_date.border = BORDER; c_date.alignment = AL_C
        tot(ws, r, 9, f"=DATEDIF(TODAY(),H{r},\"D\")", "#,##0")
        ws.cell(r, 9).font = fn(C_NAVY, bold=True, s=12)
        r += 1
    r += 1

    # ── 2027년 자금 계획 ────────────────────────────────
    sub(ws, r, 2, "2027년 자금 계획 요약", ncols=8); r += 1
    hdr(ws, r, 2, "구분"); hdr(ws, r, 3, "항목", ncols=3)
    hdr(ws, r, 6, "금액 (원)"); hdr(ws, r, 7, "비고", ncols=2); r += 1

    plan_out = [
        ("지출", "잔금 자기자금",    f"='부동산계획'!D{RE_SUMMARY_ROWS['잔금 자기자금']}"),
        ("지출", "취득세·부대비용",  f"='부동산계획'!D{RE_SUMMARY_ROWS['취득세 합계']}"),
        ("지출", "세입자 보증금 반환",f"='부동산계획'!D{RE_SUMMARY_ROWS['세입자 보증금']}"),
        ("지출", "인테리어·가전",    f"='부동산계획'!D{RE_SUMMARY_ROWS['인테리어 합계']}"),
        ("지출", "결혼식·신혼여행",  f"='결혼계획'!D{W_SUMMARY_ROWS['결혼 총비용']}"),
        ("지출", "어머니 차용금 상환","=106000000"),
    ]
    out_rows = []
    for ptype, pname, pf in plan_out:
        lbl(ws, r, 2, ptype); lbl(ws, r, 3, pname, ncols=3)
        auto(ws, r, 6, pf); inpl(ws, r, 7, ncols=2)
        out_rows.append(r); r += 1
    lbl(ws, r, 2, ""); lbl(ws, r, 3, "지출 합계", ncols=3)
    tot(ws, r, 6, f"=SUM({','.join(f'F{x}' for x in out_rows)})"); r += 1

    plan_in = [
        ("가용", "현재 가용 자산 (현금+투자-부채)", f"='자산현황'!F{AVAIL_ROW}"),
        ("가용", "2027.9까지 저축 예상",            f"='월별저축'!G{CUM_ROW}"),
        ("가용", "대출 예정",                        f"='부동산계획'!D{RE_SUMMARY_ROWS['대출 예정']}"),
        ("가용", "예상 축의금",                      f"='결혼계획'!D{W_SUMMARY_ROWS['예상 축의금']}"),
    ]
    in_rows = []
    for ptype, pname, pf in plan_in:
        lbl(ws, r, 2, ptype); lbl(ws, r, 3, pname, ncols=3)
        auto(ws, r, 6, pf); inpl(ws, r, 7, ncols=2)
        in_rows.append(r); r += 1
    lbl(ws, r, 2, ""); lbl(ws, r, 3, "가용 합계", ncols=3)
    tot(ws, r, 6, f"=SUM({','.join(f'F{x}' for x in in_rows)})")
    avail_sum_row = r; r += 1

    lbl(ws, r, 2, ""); lbl(ws, r, 3, "여유 / 부족", ncols=3)
    SURPLUS_ROW = r
    surplus_cell = tot(ws, r, 6, f"=F{avail_sum_row}-F{avail_sum_row-1}")
    surplus_cell.fill = fl(C_GREEN); r += 2

    note(ws, r, 2, "  ※ 각 시트(부동산계획·결혼계획·자산현황·월별저축)에 값을 입력하면 자동 반영됩니다.")

    return ws


# ═══════════════════════════════════════════════════════════
# 메인
# ═══════════════════════════════════════════════════════════
def main():
    import sys
    sys.stdout.reconfigure(encoding="utf-8")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("시트 생성 중...")

    # 가계부 먼저 생성 (월별저축에서 참조)
    ws_m, m_data_start, m_data_end, m_total_col = build_household(
        wb, "민엽", "민엽_가계부")
    print("  ✓ 민엽_가계부")

    ws_h, h_data_start, h_data_end, h_total_col = build_household(
        wb, "현지", "현지_가계부")
    print("  ✓ 현지_가계부")

    ws_monthly, base_row, ev_start, ev_end, data_start, data_end, cum_row = build_monthly(
        wb, "민엽_가계부", "현지_가계부",
        m_data_start, h_data_start, m_total_col, h_total_col)
    print("  ✓ 월별저축")

    ws_hist = build_history(wb)
    print("  ✓ 자산_이력")

    ws_re, re_summary_rows, deposit_row = build_realestate(wb)
    print("  ✓ 부동산계획")

    ws_w, w_summary_rows, wedding_total, trip_total, gift_total = build_wedding(wb)
    print("  ✓ 결혼계획")

    ws_assets, cash_row, inv_row, cont_row, debt_row, net_row, avail_row = build_assets(wb)
    print("  ✓ 자산현황")

    # asset_rows 튜플: (CASH_ROW, INV_ROW, CONT_ROW, DEBT_ROW, NET_ROW, AVAIL_ROW)
    asset_rows = (cash_row, inv_row, cont_row, debt_row, net_row, avail_row, avail_row)
    monthly_rows_info = (base_row, ev_start, ev_end, data_start, data_end, cum_row)

    ws_dash = build_dashboard(
        wb,
        asset_rows,
        monthly_rows_info,
        (re_summary_rows, deposit_row),
        (w_summary_rows, wedding_total, trip_total, gift_total),
    )
    print("  ✓ 대시보드")

    # 시트 순서 정렬
    order = ["대시보드","자산현황","자산_이력","민엽_가계부","현지_가계부",
             "월별저축","부동산계획","결혼계획"]
    for i, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    wb.save(OUT_PATH)
    print(f"\n저장 완료: {OUT_PATH}")
    print("시트:", wb.sheetnames)


if __name__ == "__main__":
    main()
