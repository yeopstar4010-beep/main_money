"""
patch_v3.py  —  자산계획_v3.xlsx 수정 패치
적용 항목:
  1. 결혼계획 헤더: 2027년 8월 → 2027년 7월
  2. 대시보드 D-Day 레이블 + 날짜 셀 수정
  3. 월별저축 이벤트: 잔금/취득세 날짜 2026-09 → 2027-09
  4. 자산현황 부채명: 어머니 차용금 → 대출(어머니)
  5. 자산현황 부채: 주담대·보증금반환·신용 항목 레이블 추가 (값 0)
  6. 주택청약 이동: 자산현황 현금 섹션 → 투자 섹션 연금 영역
     (레이블만 이동; 실제 금액은 파일에서 읽어 반영)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime, sys
sys.stdout.reconfigure(encoding="utf-8")

PATH = r"G:\내 드라이브\claude\자산관리\자산계획_v3.xlsx"
wb = openpyxl.load_workbook(PATH)

C_NAVY, C_BLUE, C_LBLUE = "1F3864", "2E75B6", "BDD7EE"
C_YELLOW, C_GREEN, C_RED = "FFF2CC", "E2EFDA", "FFE7E7"
C_GRAY, C_WHITE, C_DGRAY = "F2F2F2", "FFFFFF", "595959"
NUM = "#,##0"

def fl(c): return PatternFill("solid", fgColor=c)
def fn(c=C_NAVY, b=False, s=10): return Font(name="맑은 고딕", color=c, bold=b, size=s)
def bd(): s = Side(style="thin", color="BFBFBF"); return Border(left=s, right=s, top=s, bottom=s)
AL_C = Alignment(horizontal="center", vertical="center")
AL_L = Alignment(horizontal="left",   vertical="center")
AL_R = Alignment(horizontal="right",  vertical="center")
BORDER = bd()

# ── 1. 결혼계획 헤더 ────────────────────────────────────
ws = wb["결혼계획"]
for row in ws.iter_rows():
    for c in row:
        if c.value and isinstance(c.value, str) and "2027년 8월" in c.value and "결혼식" in c.value:
            c.value = c.value.replace("2027년 8월", "2027년 7월")
            print(f"  [1] 결혼계획 헤더 수정: {c.coordinate}")

# ── 2. 대시보드 D-Day ────────────────────────────────────
ws = wb["대시보드"]
for row in ws.iter_rows():
    for c in row:
        if c.value and isinstance(c.value, str) and "결혼식 (2027.08)" in c.value:
            c.value = c.value.replace("결혼식 (2027.08)", "결혼식 (2027.07)")
            print(f"  [2] 대시보드 레이블 수정: {c.coordinate}")
        if c.value == datetime.date(2027, 8, 1):
            c.value = datetime.date(2027, 7, 1)
            print(f"  [2] 대시보드 날짜 수정: {c.coordinate}")

# ── 3. 월별저축 이벤트 날짜 ─────────────────────────────
ws = wb["월별저축"]
for row in ws.iter_rows(min_col=2, max_col=4):
    ym_c, name_c = row[0], row[1]
    if ym_c.value == "2026-09":
        name = str(name_c.value or "")
        if "잔금" in name or "취득세" in name:
            ym_c.value = "2027-09"
            print(f"  [3] 월별저축 이벤트 날짜 수정: {ym_c.coordinate} ({name})")

# ── 4. 자산현황 부채명 + 추가 항목 ──────────────────────
ws = wb["자산현황"]

# 4-a. "어머니 차용금" 레이블 찾기
target_row = None
for row in ws.iter_rows(min_col=2, max_col=2):
    c = row[0]
    if c.value and "어머니 차용금" in str(c.value):
        c.value = "대출(어머니)"
        target_row = c.row
        print(f"  [4] 부채명 수정: {c.coordinate}")

# 4-b. 주택청약 처리 ─ 현금 섹션에서 찾아 0으로, 투자 섹션에서 표시
청약_민엽_val = 0
청약_현지_val = 0

for row in ws.iter_rows(min_col=2, max_col=7):
    label_c = row[0]  # D col = index 2 → col B
    if label_c.value and "주택청약" in str(label_c.value):
        # D(민엽), E(현지) 값 읽기
        d_c = row[2]  # col D
        e_c = row[3]  # col E
        if d_c.value and isinstance(d_c.value, (int, float)) and d_c.value > 0:
            청약_민엽_val += d_c.value
            d_c.value = 0
            print(f"  [5] 주택청약 민엽 값 이동: {d_c.coordinate} ({청약_민엽_val:,.0f})")
        if e_c.value and isinstance(e_c.value, (int, float)) and e_c.value > 0:
            청약_현지_val += e_c.value
            e_c.value = 0
            print(f"  [5] 주택청약 현지 값 이동: {e_c.coordinate} ({청약_현지_val:,.0f})")

# 투자 섹션에서 "IRP·연금저축"(민엽) / "연금저축"(현지) 행을 찾아 청약 값 합산
for row in ws.iter_rows(min_col=2, max_col=7):
    label_c = row[0]
    if label_c.value and ("IRP" in str(label_c.value) or "연금저축" in str(label_c.value)):
        d_c = row[2]; e_c = row[3]
        # 레이블을 청약포함으로 변경
        old = label_c.value
        if "IRP" in str(old) and 청약_민엽_val > 0:
            label_c.value = "IRP·연금·청약(민엽)"
            d_c.value = (d_c.value or 0) + 청약_민엽_val
            print(f"  [5] 청약 민엽 {청약_민엽_val:,.0f} 합산 → {label_c.coordinate}")
        if "연금저축" in str(old) and not "IRP" in str(old) and 청약_현지_val > 0:
            label_c.value = "연금·청약(현지)"
            e_c.value = (e_c.value or 0) + 청약_현지_val
            print(f"  [5] 청약 현지 {청약_현지_val:,.0f} 합산 → {e_c.coordinate}")

# 4-c. 부채 섹션에 행 삽입 (대출(주담대) / 대출(보증금 반환) / 대출(신용))
# 삽입 대신, 기존 부채 소계 행 찾기 → 소계 위에 3행 추가
if target_row:
    # 소계 행 = target_row 이후 가장 가까운 "소  계" 행
    subtotal_row = None
    for r in range(target_row + 1, target_row + 10):
        v = ws.cell(r, 2).value
        if v and "소" in str(v) and "계" in str(v):
            subtotal_row = r
            break

    if subtotal_row:
        # 3행 삽입 (subtotal_row 앞)
        ws.insert_rows(subtotal_row, 3)
        print(f"  [4] 부채 섹션에 3행 삽입 (행 {subtotal_row})")

        new_debts = [
            ("대출(주담대)",       0, "주택담보대출 실행 후 입력"),
            ("대출(보증금 반환)",  0, "보증금 반환 대출 실행 후 입력"),
            ("대출(신용)",         0, ""),
        ]
        for i, (name, val, note) in enumerate(new_debts):
            r_new = subtotal_row + i
            # B: 레이블
            lc = ws.cell(r_new, 2, name)
            lc.fill = fl(C_GRAY); lc.font = fn(C_DGRAY); lc.border = BORDER
            lc.alignment = AL_L
            # C: 빈 칸
            ws.cell(r_new, 3).border = BORDER
            # D(민엽)
            dc = ws.cell(r_new, 4, val if val != 0 else None)
            dc.fill = fl(C_WHITE); dc.font = fn(); dc.border = BORDER
            dc.alignment = AL_R; dc.number_format = NUM
            # E(현지)
            ec = ws.cell(r_new, 5, None)
            ec.fill = fl(C_WHITE); ec.font = fn(); ec.border = BORDER
            ec.alignment = AL_R; ec.number_format = NUM
            # F(합계)
            fc = ws.cell(r_new, 6, f"=D{r_new}+E{r_new}")
            fc.fill = fl(C_WHITE); fc.font = fn(); fc.border = BORDER
            fc.alignment = AL_R; fc.number_format = NUM
            # G(비고)
            gc = ws.cell(r_new, 7, note)
            gc.fill = fl(C_WHITE); gc.font = fn(); gc.border = BORDER
            gc.alignment = AL_L
            print(f"  [4] 부채 항목 추가: {name} (행 {r_new})")

        # 소계 수식 업데이트 (삽입 후 행 번호가 +3 됨)
        new_subtotal = subtotal_row + 3
        # 부채 행들: target_row ~ new_subtotal-1
        debt_rows = list(range(target_row, new_subtotal))
        for col_letter in ["D", "E"]:
            sc = ws.cell(new_subtotal, {"D":4,"E":5}[col_letter])
            sc.value = f"=SUM({','.join(f'{col_letter}{x}' for x in debt_rows)})"
        fc_sub = ws.cell(new_subtotal, 6)
        fc_sub.value = f"=D{new_subtotal}+E{new_subtotal}"
        print(f"  [4] 소계 수식 갱신: 행 {new_subtotal}")

wb.save(PATH)
print("\n패치 완료! 파일 저장됨.")
print(f"  경로: {PATH}")
print("\n수동 확인 필요:")
print("  - 순자산 합계 수식이 새 부채 행을 포함하는지 확인")
print("  - 민엽_가계부에서 '어머니 50만' 수동 삭제 필요")
