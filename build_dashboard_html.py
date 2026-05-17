"""
build_dashboard_html.py
엑셀 데이터를 읽어 dashboard.html 생성
실행: python build_dashboard_html.py
"""
import openpyxl, json, sys
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")

EXCEL_PATH = r"G:\내 드라이브\claude\자산관리\자산계획_v3.xlsx"
OUT_PATH   = r"G:\내 드라이브\claude\자산관리\dashboard.html"


# ══════════════════════════════════════════════════════════
# 데이터 로드
# ══════════════════════════════════════════════════════════
def load():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    def g(ws, r, c, default=0):
        v = ws.cell(r, c).value
        return v if isinstance(v, (int, float)) else default

    d = {}

    # ── 자산현황 (formula 셀은 None → 직접 합산) ──────────
    ws = wb["자산현황"]
    d["현금_m"]   = sum(g(ws, r, 3) for r in range(6, 16))
    d["현금_h"]   = sum(g(ws, r, 5) for r in range(6, 16))
    d["현금_t"]   = d["현금_m"] + d["현금_h"]
    d["투자_m"]   = sum(g(ws, r, 3) for r in range(19, 25))
    d["투자_h"]   = sum(g(ws, r, 5) for r in range(19, 25))
    d["투자_t"]   = d["투자_m"] + d["투자_h"]
    d["계약금_m"] = g(ws, 28, 3)
    d["계약금_t"] = d["계약금_m"]
    d["부채_m"]   = sum(g(ws, r, 3) for r in range(32, 36))
    d["부채_h"]   = sum(g(ws, r, 5) for r in range(32, 36))
    d["부채_t"]   = d["부채_m"] + d["부채_h"]
    d["순자산_m"] = d["현금_m"] + d["투자_m"] + d["계약금_m"] + d["부채_m"]
    d["순자산_h"] = d["현금_h"] + d["투자_h"] + d["부채_h"]
    d["순자산_t"] = d["순자산_m"] + d["순자산_h"]
    d["가용_t"]   = d["현금_t"] + d["투자_t"] + d["부채_t"]

    # 자산 상세 (도넛용)
    inv_rows_m = [(ws.cell(r,2).value, g(ws,r,3)) for r in range(19,25) if g(ws,r,3)>0]
    inv_rows_h = [(ws.cell(r,4).value, g(ws,r,5)) for r in range(19,25) if g(ws,r,5)>0]
    d["inv_m"] = inv_rows_m
    d["inv_h"] = inv_rows_h

    # ── 부동산계획 ─────────────────────────────────────────
    ws = wb["부동산계획"]
    d["매매가"]   = g(ws, 5, 4)
    d["잔금_자"]  = g(ws, 7, 4, 220_000_000)
    d["대출"]     = g(ws, 8, 4, 560_000_000)
    d["취득세"]   = g(ws, 9, 4, 50_000_000)
    d["보증금"]   = g(ws, 10, 4, 300_000_000)
    d["인테리어"] = g(ws, 11, 4, 40_000_000)

    # ── 결혼계획 ───────────────────────────────────────────
    ws = wb["결혼계획"]
    d["결혼식"]  = g(ws, 5, 4, 50_000_000)
    d["신혼여행"] = g(ws, 6, 4, 10_000_000)
    d["결혼총"]  = g(ws, 7, 4) or d["결혼식"] + d["신혼여행"]
    d["축의금"]  = g(ws, 8, 4, 30_000_000)

    # ── 월별저축 ───────────────────────────────────────────
    ws = wb["월별저축"]
    monthly = []
    for r in range(21, 21 + 33):
        bval = ws.cell(r, 2).value
        if not isinstance(bval, (datetime, date)):
            break
        if isinstance(bval, datetime):
            bval = bval.date()
        m_in  = g(ws, r, 3)
        h_in  = g(ws, r, 4)
        m_out = g(ws, r, 6)
        h_out = g(ws, r, 7)
        t_in  = m_in + h_in
        t_out = m_out + h_out
        monthly.append({
            "ym":       bval.strftime("%y.%m"),
            "ym_full":  bval.strftime("%Y-%m"),
            "민엽수입": m_in,
            "현지수입": h_in,
            "합계수입": t_in,
            "합계지출": t_out,
            "저축액":   t_in - t_out,
        })

    # 2027.09까지만 누계
    target = [m for m in monthly if m["ym_full"] <= "2027-09"]
    d["누계저축"] = sum(m["저축액"] for m in target)
    d["월별"]     = monthly

    return d


# ══════════════════════════════════════════════════════════
# 포맷 헬퍼
# ══════════════════════════════════════════════════════════
def fmt(n, sign=False):
    """숫자 → 억/만 한국식"""
    if n is None or n == 0:
        return "0"
    neg = n < 0
    n = abs(n)
    if n >= 100_000_000:
        eok = n // 100_000_000
        man = (n % 100_000_000) // 10_000
        s = f"{eok}억" + (f" {man:,}만" if man else "")
    elif n >= 10_000:
        s = f"{n // 10_000:,}만"
    else:
        s = f"{n:,}"
    prefix = "△" if neg else ("+" if sign else "")
    return prefix + s

def fmt_m(n):
    """만원 단위 정수"""
    return f"{int(abs(n) / 10_000):,}만"


# ══════════════════════════════════════════════════════════
# HTML 생성
# ══════════════════════════════════════════════════════════
def build_html(d):
    today = date.today()
    wedding_day = date(2027, 7, 1)
    balance_day = date(2027, 9, 1)
    d_wed = (wedding_day - today).days
    d_bal = (balance_day - today).days

    # D-Day 진행도 (2024-01-01 기준)
    total_days = (wedding_day - date(2024, 1, 1)).days
    elapsed    = (today - date(2024, 1, 1)).days
    prog_wed   = min(100, int(elapsed / total_days * 100))
    total_b    = (balance_day - date(2024, 1, 1)).days
    prog_bal   = min(100, int(elapsed / total_b * 100))

    # 자금 계획 수치
    총지출 = (abs(d["잔금_자"]) + abs(d["취득세"]) + abs(d["보증금"])
             + abs(d["인테리어"]) + abs(d["결혼총"]) + 106_000_000)
    총가용 = d["가용_t"] + d["누계저축"] + d["대출"] + d["축의금"]
    여유   = 총가용 - 총지출

    # 차트 데이터 JSON
    # 1) 자산 구성 도넛 (합계)
    donut_labels = ["현금·예금", "투자자산", "계약금"]
    donut_values = [max(0, d["현금_t"]), max(0, d["투자_t"]), max(0, d["계약금_t"])]

    # 2) 민엽/현지 비교 bar
    bar_cats   = ["현금·예금", "투자자산", "계약금", "부채"]
    bar_minyeob = [d["현금_m"], d["투자_m"], d["계약금_m"], d["부채_m"]]
    bar_hyeonji = [d["현금_h"], d["투자_h"], 0,            d["부채_h"]]

    # 3) 자금 계획 horizontal bar
    fund_out = [
        ("잔금 자기자금",     abs(d["잔금_자"])),
        ("취득세·부대비용",   abs(d["취득세"])),
        ("세입자 보증금 반환",abs(d["보증금"])),
        ("인테리어·가전",     abs(d["인테리어"])),
        ("결혼식·신혼여행",   abs(d["결혼총"])),
        ("어머니 차용금",     106_000_000),
    ]
    fund_in = [
        ("현재 가용자산",     d["가용_t"]),
        ("2027.09 저축 예상", d["누계저축"]),
        ("대출 예정",         d["대출"]),
        ("예상 축의금",       d["축의금"]),
    ]

    # 4) 월별 추이 (2027.09까지)
    target_monthly = [m for m in d["월별"] if m["ym_full"] <= "2027-09"]
    months_ym   = [m["ym"] for m in target_monthly]
    수입_arr    = [m["합계수입"] for m in target_monthly]
    지출_arr    = [m["합계지출"] for m in target_monthly]
    저축_arr    = [m["저축액"] for m in target_monthly]
    누계_arr    = []
    acc = 0
    for m in target_monthly:
        acc += m["저축액"]
        누계_arr.append(acc)

    def j(v): return json.dumps(v, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>민엽 & 현지 · 자산관리</title>
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link rel="stylesheet" as="style" crossorigin href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css">
  <script src="https://cdn.jsdelivr.net/npm/apexcharts"></script>
  <style>
    :root {{
      --navy:    #1F3864;
      --blue:    #2E75B6;
      --blue2:   #5BA3D9;
      --lblue:   #BDD7EE;
      --bg:      #F4F7FB;
      --card:    #FFFFFF;
      --border:  #E6EDF7;
      --text:    #1A2332;
      --muted:   #6B7A8D;
      --green:   #10B981;
      --red:     #EF4444;
      --r:       20px;
      --r-sm:    12px;
      --shadow:  0 1px 3px rgba(31,56,100,.05), 0 4px 20px rgba(31,56,100,.07);
      --shadow-h:0 8px 40px rgba(31,56,100,.14);
    }}
    *, *::before, *::after {{
      box-sizing: border-box;
      font-family: 'Pretendard', -apple-system, sans-serif;
    }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--text);
      min-height: 100vh;
    }}
    /* ── layout ── */
    .wrap {{
      max-width: 1440px;
      margin: 0 auto;
      padding: 40px 48px 64px;
    }}
    /* ── header ── */
    .hd {{
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      padding-bottom: 28px;
      margin-bottom: 32px;
      border-bottom: 1px solid var(--border);
    }}
    .hd-title   {{ font-size: 28px; font-weight: 800; color: var(--navy); letter-spacing: -.5px; }}
    .hd-sub     {{ font-size: 13px; color: var(--muted); margin-top: 5px; font-weight: 400; }}
    .hd-date    {{ font-size: 13px; color: var(--muted); text-align: right; line-height: 1.8; }}
    .hd-date b  {{ color: var(--navy); font-weight: 600; }}
    /* ── card ── */
    .card {{
      background: var(--card);
      border-radius: var(--r);
      border: 1px solid var(--border);
      box-shadow: var(--shadow);
      padding: 24px 28px;
      transition: box-shadow .2s;
    }}
    .card:hover {{ box-shadow: var(--shadow-h); }}
    .card-sm {{ padding: 18px 22px; border-radius: 16px; }}
    /* ── gradient card ── */
    .card-navy {{
      background: linear-gradient(135deg, #1B3260 0%, #2E75B6 100%);
      color: white;
      border-color: transparent;
    }}
    /* ── metric ── */
    .lbl {{
      font-size: 11px; font-weight: 600; letter-spacing: .8px;
      text-transform: uppercase; color: var(--muted); margin-bottom: 8px;
    }}
    .card-navy .lbl {{ color: rgba(255,255,255,.65); }}
    .val {{
      font-size: 28px; font-weight: 800; color: var(--navy);
      line-height: 1; letter-spacing: -.5px;
    }}
    .card-navy .val {{ color: #fff; }}
    .sub {{
      font-size: 12px; color: var(--muted); margin-top: 8px; line-height: 1.6;
    }}
    .card-navy .sub {{ color: rgba(255,255,255,.65); }}
    .split {{
      display: flex; gap: 20px; margin-top: 14px;
      padding-top: 14px; border-top: 1px solid rgba(255,255,255,.15);
    }}
    .split-item .split-lbl {{ font-size: 10px; color: rgba(255,255,255,.6); margin-bottom: 2px; }}
    .split-item .split-val {{ font-size: 15px; font-weight: 700; color: #fff; }}
    .vr {{ width: 1px; background: rgba(255,255,255,.2); }}
    /* ── row & col ── */
    .row {{ display: flex; gap: 16px; margin-bottom: 16px; }}
    .row > .card {{ flex: 1; }}
    /* ── D-Day ── */
    .dday {{ text-align: center; padding: 20px 16px; }}
    .dday-ev {{ font-size: 12px; font-weight: 600; color: var(--blue); margin-bottom: 10px; letter-spacing: .3px; }}
    .dday-num {{ font-size: 38px; font-weight: 900; color: var(--navy); letter-spacing: -2px; line-height: 1; }}
    .dday-pref {{ font-size: 15px; font-weight: 700; color: var(--blue2); }}
    .dday-dt {{ font-size: 11px; color: var(--muted); margin-top: 8px; }}
    .prog-bg {{ height: 4px; background: var(--border); border-radius: 99px; margin-top: 14px; }}
    .prog-fill {{ height: 4px; border-radius: 99px; background: linear-gradient(90deg, var(--blue), var(--blue2)); }}
    /* ── section title ── */
    .sect {{
      font-size: 15px; font-weight: 700; color: var(--navy);
      margin-bottom: 16px; display: flex; align-items: center; gap: 8px;
    }}
    .sect::before {{
      content: ""; display: block; width: 3px; height: 15px;
      border-radius: 99px;
      background: linear-gradient(180deg, var(--blue), var(--navy));
    }}
    /* ── chart card ── */
    .chart-card {{ padding: 28px 28px 16px; }}
    /* ── fund list ── */
    .fund-item {{
      display: flex; align-items: center; gap: 10px;
      padding: 9px 0; border-bottom: 1px solid #F0F5FD;
    }}
    .fund-item:last-child {{ border-bottom: none; }}
    .dot {{ width: 7px; height: 7px; border-radius: 50%; flex-shrink: 0; }}
    .dot-out {{ background: var(--red); }}
    .dot-in  {{ background: var(--blue); }}
    .fund-name {{ flex: 1; font-size: 13px; color: var(--text); }}
    .fund-amt  {{ font-size: 13px; font-weight: 600; }}
    .fund-amt.neg {{ color: var(--red); }}
    .fund-amt.pos {{ color: var(--blue); }}
    .fund-divider {{
      margin: 8px 0; border: none;
      border-top: 1px dashed #D6E4F5; height: 1px;
    }}
    /* ── summary chips ── */
    .chip-row {{ display: flex; gap: 10px; margin-top: 18px; }}
    .chip {{
      flex: 1; background: #F0F6FF; border-radius: 12px;
      padding: 12px 16px; text-align: center;
    }}
    .chip.red  {{ background: #FFF1F2; }}
    .chip.green{{ background: #F0FDF4; }}
    .chip-lbl  {{ font-size: 10px; color: var(--muted); font-weight: 600;
                  letter-spacing: .6px; text-transform: uppercase; margin-bottom: 4px; }}
    .chip-val  {{ font-size: 17px; font-weight: 800; color: var(--navy); }}
    .chip.red  .chip-val {{ color: var(--red); }}
    .chip.green .chip-val {{ color: var(--green); }}
    /* ── table ── */
    .mini-table {{ width: 100%; font-size: 13px; border-collapse: collapse; }}
    .mini-table td {{ padding: 8px 4px; border-bottom: 1px solid #F0F5FD; }}
    .mini-table tr:last-child td {{ border-bottom: none; }}
    .mini-table .td-lbl {{ color: var(--muted); font-size: 12px; }}
    .mini-table .td-val {{ text-align: right; font-weight: 600; color: var(--navy); }}
    /* ── footer ── */
    .footer {{
      text-align: center; margin-top: 48px;
      font-size: 11px; color: #B0BFCC;
    }}
    /* ── 반응형 ── */
    @media (max-width: 900px) {{
      .wrap {{ padding: 20px 20px 40px; }}
      .row {{ flex-wrap: wrap; }}
      .hd {{ flex-direction: column; gap: 8px; align-items: flex-start; }}
    }}
  </style>
</head>
<body>
<div class="wrap">

  <!-- ── header ── -->
  <div class="hd">
    <div>
      <div class="hd-title">민엽 & 현지 · 자산관리</div>
      <div class="hd-sub">2027년 결혼·입주를 위한 통합 자금 계획 대시보드</div>
    </div>
    <div class="hd-date">
      기준일 &nbsp;<b>{today.strftime('%Y. %m. %d')}</b><br>
      <span style="font-size:11px">데이터: 자산계획_v3.xlsx</span>
    </div>
  </div>

  <!-- ── row 1: 순자산 + 현금 + 투자 + D-Day ── -->
  <div class="row" style="align-items:stretch;">

    <!-- 순자산 그라디언트 카드 -->
    <div class="card card-navy" style="flex:1.7; min-width:0;">
      <div class="lbl">순자산 합계</div>
      <div class="val" style="font-size:34px;">{fmt(d["순자산_t"])}</div>
      <div class="sub">가용자산 &nbsp;{fmt(d["가용_t"])}</div>
      <div class="split">
        <div class="split-item">
          <div class="split-lbl">민 엽</div>
          <div class="split-val">{fmt(d["순자산_m"])}</div>
        </div>
        <div class="vr"></div>
        <div class="split-item">
          <div class="split-lbl">현 지</div>
          <div class="split-val">{fmt(d["순자산_h"])}</div>
        </div>
      </div>
    </div>

    <!-- 현금·예금 -->
    <div class="card" style="flex:1; min-width:0;">
      <div class="lbl">현금 · 예금</div>
      <div class="val" style="font-size:22px;">{fmt(d["현금_t"])}</div>
      <div class="sub">
        민엽 {fmt(d["현금_m"])}<br>현지 {fmt(d["현금_h"])}
      </div>
    </div>

    <!-- 투자자산 -->
    <div class="card" style="flex:1; min-width:0;">
      <div class="lbl">투자 자산</div>
      <div class="val" style="font-size:22px;">{fmt(d["투자_t"])}</div>
      <div class="sub">
        민엽 {fmt(d["투자_m"])}<br>현지 {fmt(d["투자_h"])}
      </div>
    </div>

    <!-- D-Day 결혼식 -->
    <div class="card dday" style="flex:.85; min-width:0;">
      <div class="dday-ev">결혼식</div>
      <div class="dday-num"><span class="dday-pref">D-</span>{d_wed:,}</div>
      <div class="dday-dt">2027년 7월</div>
      <div class="prog-bg">
        <div class="prog-fill" style="width:{prog_wed}%;"></div>
      </div>
    </div>

    <!-- D-Day 잔금 -->
    <div class="card dday" style="flex:.85; min-width:0;">
      <div class="dday-ev">잔금 · 입주</div>
      <div class="dday-num"><span class="dday-pref">D-</span>{d_bal:,}</div>
      <div class="dday-dt">2027년 9월</div>
      <div class="prog-bg">
        <div class="prog-fill" style="width:{prog_bal}%;"></div>
      </div>
    </div>

  </div><!-- /row1 -->

  <!-- ── row 2: 자산 구성 | 2027 자금 계획 ── -->
  <div class="row" style="align-items:stretch;">

    <!-- 자산 구성 -->
    <div class="card chart-card" style="flex:1.1; min-width:0;">
      <div class="sect">자산 구성</div>
      <div id="donut" style="margin: 0 auto;"></div>
      <div style="margin-top:20px;">
        <div class="sect" style="margin-bottom:12px;">민엽 · 현지 비교</div>
        <div id="bar-compare" style="min-height:180px;"></div>
      </div>
    </div>

    <!-- 2027 자금 계획 -->
    <div class="card chart-card" style="flex:1; min-width:0;">
      <div class="sect">2027년 자금 계획</div>
      <div id="fund-chart" style="min-height:320px;"></div>
      <div class="chip-row">
        <div class="chip red">
          <div class="chip-lbl">지출 합계</div>
          <div class="chip-val">{fmt(총지출)}</div>
        </div>
        <div class="chip">
          <div class="chip-lbl">가용 합계</div>
          <div class="chip-val" style="color:var(--blue);">{fmt(총가용)}</div>
        </div>
        <div class="chip {'green' if 여유 >= 0 else 'red'}">
          <div class="chip-lbl">{'여유' if 여유 >= 0 else '부족'}</div>
          <div class="chip-val">{fmt(abs(여유))}</div>
        </div>
      </div>
    </div>

  </div><!-- /row2 -->

  <!-- ── row 3: 월별 저축 추이 ── -->
  <div class="card chart-card" style="margin-bottom:16px;">
    <div class="sect">월별 저축 추이 &nbsp;<span style="font-size:12px;font-weight:400;color:var(--muted);">2026.05 → 2027.09</span></div>
    <div id="monthly-chart" style="min-height:300px;"></div>
    <!-- 요약 칩 -->
    <div class="chip-row" style="margin-top:12px;">
      <div class="chip">
        <div class="chip-lbl">2027.09까지 예상 저축</div>
        <div class="chip-val" style="color:var(--blue);">{fmt(d["누계저축"])}</div>
      </div>
      <div class="chip">
        <div class="chip-lbl">남은 기간</div>
        <div class="chip-val">{d_bal // 30}개월</div>
      </div>
      <div class="chip">
        <div class="chip-lbl">이미 납부 계약금</div>
        <div class="chip-val">{fmt(d["계약금_t"])}</div>
      </div>
      <div class="chip">
        <div class="chip-lbl">어머니 차용금 (부채)</div>
        <div class="chip-val" style="color:var(--red);">△{fmt(abs(d["부채_t"]))}</div>
      </div>
    </div>
  </div><!-- /row3 -->

  <!-- ── footer ── -->
  <div class="footer">
    민엽 &amp; 현지 자산관리 &nbsp;·&nbsp; {today.strftime('%Y. %m. %d')} 기준
    &nbsp;·&nbsp; 데이터: 자산계획_v3.xlsx
  </div>

</div><!-- /wrap -->

<script>
/* ── 공통 팔레트 ── */
const NAVY   = "#1F3864";
const BLUE   = "#2E75B6";
const BLUE2  = "#5BA3D9";
const LBLUE  = "#BDD7EE";
const RED    = "#EF4444";
const GREEN  = "#10B981";
const MUTED  = "#6B7A8D";
const FONT   = "Pretendard, -apple-system, sans-serif";
const baseOpts = {{
  chart:  {{ fontFamily: FONT, toolbar: {{ show: false }}, animations: {{ speed: 600 }} }},
  tooltip:{{ style: {{ fontFamily: FONT, fontSize: "12px" }} }},
  legend: {{ fontFamily: FONT, fontSize: "12px" }},
  dataLabels: {{ style: {{ fontFamily: FONT }} }},
}};

/* ══════════════════════════════════════════
   1. 자산 구성 도넛
══════════════════════════════════════════ */
new ApexCharts(document.getElementById("donut"), {{
  ...baseOpts,
  chart: {{ ...baseOpts.chart, type: "donut", height: 280 }},
  series: {j(donut_values)},
  labels: {j(donut_labels)},
  colors: [LBLUE, BLUE, NAVY],
  plotOptions: {{
    pie: {{
      donut: {{
        size: "68%",
        labels: {{
          show: true,
          name:  {{ fontFamily: FONT, fontSize: "13px", color: NAVY }},
          value: {{ fontFamily: FONT, fontSize: "18px", fontWeight: 800, color: NAVY,
                    formatter: v => numFmt(v) }},
          total: {{
            show: true, label: "총 자산", fontFamily: FONT, fontSize: "12px", color: MUTED,
            formatter: w => numFmt(w.globals.seriesTotals.reduce((a,b)=>a+b,0)),
          }},
        }},
      }},
    }},
  }},
  dataLabels: {{ enabled: false }},
  legend: {{ ...baseOpts.legend, position: "bottom", markers: {{ radius: 4 }} }},
  tooltip: {{ ...baseOpts.tooltip, y: {{ formatter: v => numFmt(v) + "원" }} }},
}}).render();

/* ══════════════════════════════════════════
   2. 민엽/현지 비교 bar
══════════════════════════════════════════ */
new ApexCharts(document.getElementById("bar-compare"), {{
  ...baseOpts,
  chart:  {{ ...baseOpts.chart, type: "bar", height: 190, stacked: false }},
  series: [
    {{ name: "민엽", data: {j(bar_minyeob)} }},
    {{ name: "현지", data: {j(bar_hyeonji)} }},
  ],
  xaxis:  {{ categories: {j(bar_cats)}, labels: {{ style: {{ fontFamily: FONT, fontSize: "11px" }} }} }},
  yaxis:  {{ labels: {{ formatter: v => numFmt(v), style: {{ fontFamily: FONT, fontSize: "10px" }} }} }},
  colors: [NAVY, BLUE2],
  plotOptions: {{ bar: {{ borderRadius: 4, columnWidth: "60%", borderRadiusApplication: "end" }} }},
  dataLabels: {{ enabled: false }},
  grid: {{ borderColor: "#F0F5FD", xaxis: {{ lines: {{ show: false }} }} }},
  legend: {{ ...baseOpts.legend, position: "top", horizontalAlign: "right" }},
  tooltip: {{ ...baseOpts.tooltip, y: {{ formatter: v => numFmt(v) + "원" }} }},
}}).render();

/* ══════════════════════════════════════════
   3. 자금 계획 가로 bar
══════════════════════════════════════════ */
const fundLabels = {j([n for n,_ in fund_out + fund_in])};
const fundValues = {j([v for _,v in fund_out + fund_in])};
const fundColors = {j(["#EF4444"]*len(fund_out) + ["#2E75B6"]*len(fund_in))};
const fundOpacity= {j([0.85]*len(fund_out) + [0.85]*len(fund_in))};

new ApexCharts(document.getElementById("fund-chart"), {{
  ...baseOpts,
  chart:  {{ ...baseOpts.chart, type: "bar", height: 340 }},
  series: [{{ name: "금액", data: fundValues }}],
  xaxis: {{
    categories: fundLabels,
    labels: {{ show: false }},
  }},
  yaxis: {{
    labels: {{
      style: {{ fontFamily: FONT, fontSize: "11.5px", colors: Array(fundLabels.length).fill(MUTED) }},
      maxWidth: 160,
    }},
  }},
  plotOptions: {{
    bar: {{
      horizontal: true, borderRadius: 5,
      borderRadiusApplication: "end",
      dataLabels: {{ position: "top" }},
    }},
  }},
  fill: {{ colors: fundColors, opacity: fundOpacity }},
  dataLabels: {{
    enabled: true, offsetX: 6,
    style: {{ fontFamily: FONT, fontSize: "11px", colors: [NAVY], fontWeight: 600 }},
    formatter: v => numFmt(v),
  }},
  grid: {{ borderColor: "#F0F5FD", yaxis: {{ lines: {{ show: false }} }} }},
  legend: {{ show: false }},
  tooltip: {{ ...baseOpts.tooltip, y: {{ formatter: v => numFmt(v) + "원" }} }},
}}).render();

/* ══════════════════════════════════════════
   4. 월별 저축 추이
══════════════════════════════════════════ */
new ApexCharts(document.getElementById("monthly-chart"), {{
  ...baseOpts,
  chart: {{ ...baseOpts.chart, type: "line", height: 310,
           zoom: {{ enabled: false }}, toolbar: {{ show: false }} }},
  series: [
    {{ name: "합계 수입", type: "column", data: {j(수입_arr)} }},
    {{ name: "합계 지출", type: "column", data: {j(지출_arr)} }},
    {{ name: "월 저축액", type: "line",   data: {j(저축_arr)} }},
    {{ name: "누계 저축", type: "line",   data: {j(누계_arr)}, yAxisIndex: 1 }},
  ],
  xaxis: {{
    categories: {j(months_ym)},
    tickAmount: 8,
    labels: {{ style: {{ fontFamily: FONT, fontSize: "10px", colors: MUTED }} }},
  }},
  yaxis: [
    {{ seriesName: "합계 수입",
       labels: {{ formatter: v => numFmt(v), style: {{ fontFamily: FONT, fontSize: "10px", colors: [MUTED] }} }} }},
    {{ seriesName: "누계 저축", opposite: true,
       labels: {{ formatter: v => numFmt(v), style: {{ fontFamily: FONT, fontSize: "10px", colors: [BLUE] }} }} }},
  ],
  colors: [LBLUE, "#FECACA", NAVY, BLUE],
  fill: {{
    opacity: [0.7, 0.7, 1, 0.08],
    type: ["solid","solid","solid","gradient"],
    gradient: {{ shade:"light", type:"vertical", shadeIntensity:.3,
                  gradientToColors:[BLUE], opacityFrom:.12, opacityTo:.01 }},
  }},
  stroke: {{
    width: [0, 0, 2.5, 2.5],
    curve: ["smooth","smooth","smooth","smooth"],
    dashArray: [0, 0, 0, 5],
  }},
  markers: {{
    size: [0, 0, 4, 0],
    colors: [NAVY],
    strokeColors: "#fff",
    strokeWidth: 2,
  }},
  plotOptions: {{ bar: {{ columnWidth: "55%", borderRadius: 3, borderRadiusApplication: "end" }} }},
  dataLabels: {{ enabled: false }},
  legend: {{ ...baseOpts.legend, position: "top", horizontalAlign: "left",
             markers: {{ radius: 4 }}, offsetY: -4 }},
  grid: {{ borderColor: "#F0F5FD", xaxis: {{ lines: {{ show: false }} }} }},
  tooltip: {{ ...baseOpts.tooltip, shared: true, intersect: false,
              y: {{ formatter: v => v != null ? numFmt(v)+"원" : "-" }} }},
}}).render();

/* ══════════════════════════════════════════
   numFmt helper
══════════════════════════════════════════ */
function numFmt(n) {{
  if (n == null) return "-";
  const abs = Math.abs(n);
  if (abs >= 1e8) {{
    const eok = Math.floor(abs / 1e8);
    const man = Math.floor((abs % 1e8) / 1e4);
    return man > 0 ? `${{eok}}억 ${{man.toLocaleString()}}만` : `${{eok}}억`;
  }} else if (abs >= 1e4) {{
    return `${{Math.floor(abs / 1e4).toLocaleString()}}만`;
  }}
  return abs.toLocaleString();
}}
</script>

</body>
</html>"""

    # 차트 데이터 변수 주입
    data_vars = f"""
const donut_values = {j(donut_values)};
const donut_labels = {j(donut_labels)};
const bar_minyeob  = {j(bar_minyeob)};
const bar_hyeonji  = {j(bar_hyeonji)};
const bar_cats     = {j(bar_cats)};
"""
    # html에 이미 인라인으로 삽입했으므로 별도 삽입 불필요
    return html


# ══════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("엑셀 데이터 로드 중...")
    data = load()

    print("HTML 생성 중...")
    html = build_html(data)

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n완료! 파일: {OUT_PATH}")
    print(f"  순자산 합계: {fmt(data['순자산_t'])}")
    print(f"  가용자산:   {fmt(data['가용_t'])}")
    print(f"  누계저축:   {fmt(data['누계저축'])}")

    import os, subprocess
    subprocess.Popen(["start", "", OUT_PATH], shell=True)
    print("  브라우저에서 대시보드를 열었습니다.")
