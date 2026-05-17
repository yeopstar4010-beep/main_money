"""
자산관리 엑셀 시트 추가 스크립트
실행: python add_sheets.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy

EXCEL_PATH = r"G:\내 드라이브\claude\자산관리\자산계획.xlsx"

# ── 스타일 헬퍼 ──────────────────────────────────────────
H1_FILL   = PatternFill("solid", fgColor="1F3864")
H2_FILL   = PatternFill("solid", fgColor="2E75B6")
H3_FILL   = PatternFill("solid", fgColor="BDD7EE")
SUM_FILL  = PatternFill("solid", fgColor="FFF2CC")
WHITE_FILL= PatternFill("solid", fgColor="FFFFFF")

H1_FONT  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
H2_FONT  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
H3_FONT  = Font(name="맑은 고딕", bold=True, color="1F3864", size=10)
BODY_FONT= Font(name="맑은 고딕", size=10)
SUM_FONT = Font(name="맑은 고딕", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT  = '#,##0'
PCT_FMT  = '0.0%'
DATE_FMT = 'YYYY-MM'


def h1(ws, row, col, value, cols=1):
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.fill, c.alignment = H1_FONT, H1_FILL, CENTER
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + cols - 1)
    return c

def h2(ws, row, col, value, cols=1):
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.fill, c.alignment = H2_FONT, H2_FILL, CENTER
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + cols - 1)
    return c

def h3(ws, row, col, value, cols=1):
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.fill, c.alignment = H3_FONT, H3_FILL, CENTER
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + cols - 1)
    return c

def label(ws, row, col, value="", cols=1):
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.alignment, c.border = BODY_FONT, LEFT, BORDER
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + cols - 1)
    return c

def inp(ws, row, col, value=None, fmt=NUM_FMT):
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.alignment, c.border = BODY_FONT, RIGHT, BORDER
    c.number_format = fmt
    return c

def total(ws, row, col, formula, fmt=NUM_FMT, cols=1):
    c = ws.cell(row=row, column=col, value=formula)
    c.font, c.fill, c.alignment, c.border = SUM_FONT, SUM_FILL, RIGHT, BORDER
    c.number_format = fmt
    if cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + cols - 1)
    return c

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ══════════════════════════════════════════════════════════
# 1. 💒 결혼식_세부예산
# ══════════════════════════════════════════════════════════
def build_wedding(wb):
    if "💒 결혼식_세부예산" in wb.sheetnames:
        del wb["💒 결혼식_세부예산"]
    ws = wb.create_sheet("💒 결혼식_세부예산")

    # 열 너비
    ws.column_dimensions["A"].width = 2
    set_col_width(ws, 2, 20)  # 항목
    set_col_width(ws, 3, 16)  # 예산
    set_col_width(ws, 4, 16)  # 실제비용
    set_col_width(ws, 5, 16)  # 차액
    set_col_width(ws, 6, 22)  # 비고

    r = 2
    h1(ws, r, 2, "💒  결혼식 세부 예산", 5); r += 1
    ws.row_dimensions[r-1].height = 30

    # 헤더
    h3(ws, r, 2, "항  목")
    h3(ws, r, 3, "예 산 (원)")
    h3(ws, r, 4, "실제 비용 (원)")
    h3(ws, r, 5, "차  액")
    h3(ws, r, 6, "비  고")
    r += 1

    sections = [
        ("🎨 스드메", [
            ("스튜디오 (사진 촬영)", None),
            ("드레스 (웨딩드레스)", None),
            ("메이크업 (신부 헤메)", None),
            ("신랑 예복", None),
            ("부케 · 부토니에", None),
        ]),
        ("🏛 예식장", [
            ("예식장 대관료", None),
            ("음식 (뷔페 / 한식)", None),
            ("웨딩 케이크", None),
            ("사회자", None),
            ("영상 촬영", None),
            ("청첩장 · 답례품", None),
        ]),
        ("💍 예물 · 예단", [
            ("예물 (반지 등)", None),
            ("예단 비용", None),
        ]),
        ("🏠 살림 준비", [
            ("혼수 (가전 · 가구)", None),
            ("기타 생활용품", None),
        ]),
        ("기 타", [
            ("기타 ①", None),
            ("기타 ②", None),
        ]),
    ]

    section_start_rows = []
    for sec_name, items in sections:
        h2(ws, r, 2, sec_name, 5); r += 1
        sec_item_rows = []
        for item_name, _ in items:
            label(ws, r, 2, item_name)
            inp(ws, r, 3)   # 예산
            inp(ws, r, 4)   # 실제비용
            # 차액 = 실제비용 - 예산
            col_d = get_column_letter(4)
            col_c = get_column_letter(3)
            inp(ws, r, 5, f"={col_d}{r}-{col_c}{r}", NUM_FMT)
            ws.cell(r, 5).font = BODY_FONT
            label(ws, r, 6, "")  # 비고
            sec_item_rows.append(r)
            r += 1
        section_start_rows.append((sec_name, sec_item_rows))
        # 소계
        rows_ref = ",".join([f"C{x}" for x in sec_item_rows])
        total(ws, r, 2, f"  {sec_name} 소계")
        ws.cell(r, 2).alignment = LEFT
        total(ws, r, 3, f"=SUM({rows_ref})")
        rows_ref_d = ",".join([f"D{x}" for x in sec_item_rows])
        total(ws, r, 4, f"=SUM({rows_ref_d})")
        total(ws, r, 5, f"=D{r}-C{r}")
        ws.cell(r, 6).fill = SUM_FILL
        r += 1

    # 전체 합계
    r += 1
    h1(ws, r, 2, "총  합  계", 1)
    ws.cell(r, 2).alignment = LEFT
    total(ws, r, 3, "=SUM(C3:C100)")
    total(ws, r, 4, "=SUM(D3:D100)")
    total(ws, r, 5, f"=D{r}-C{r}")
    ws.row_dimensions[r].height = 22

    # 안내
    r += 2
    ws.cell(r, 2, "  C열 예산, D열 실제 비용을 입력하세요. 차액과 합계는 자동 계산됩니다.").font = Font(name="맑은 고딕", size=9, italic=True, color="808080")


# ══════════════════════════════════════════════════════════
# 2. 🎊 축의금_시뮬레이터
# ══════════════════════════════════════════════════════════
def build_gift(wb):
    if "🎊 축의금_시뮬레이터" in wb.sheetnames:
        del wb["🎊 축의금_시뮬레이터"]
    ws = wb.create_sheet("🎊 축의금_시뮬레이터")

    ws.column_dimensions["A"].width = 2
    set_col_width(ws, 2, 22)
    set_col_width(ws, 3, 14)
    set_col_width(ws, 4, 14)
    set_col_width(ws, 5, 14)
    set_col_width(ws, 6, 14)
    set_col_width(ws, 7, 18)

    r = 2
    h1(ws, r, 2, "🎊  축의금 시뮬레이터", 6); r += 1
    ws.row_dimensions[r-1].height = 30

    # 구분별 테이블
    h2(ws, r, 2, "하객 구분별 예상", 6); r += 1
    h3(ws, r, 2, "구  분")
    h3(ws, r, 3, "민엽 하객수")
    h3(ws, r, 4, "현지 하객수")
    h3(ws, r, 5, "합계 하객수")
    h3(ws, r, 6, "평균 축의금")
    h3(ws, r, 7, "예상 축의금")
    r += 1

    groups = [
        ("직장 동료",    None, None, 70000),
        ("친구 · 지인", None, None, 100000),
        ("친척 · 가족", None, None, 200000),
        ("기타",         None, None, 100000),
    ]

    group_rows = []
    for g_name, minyeob, hyunji, avg in groups:
        label(ws, r, 2, g_name)
        inp(ws, r, 3, minyeob)       # 민엽 하객수
        inp(ws, r, 4, hyunji)        # 현지 하객수
        inp(ws, r, 5, f"=C{r}+D{r}", '#,##0')  # 합계
        inp(ws, r, 6, avg)           # 평균 축의금
        total(ws, r, 7, f"=E{r}*F{r}")  # 예상
        ws.cell(r, 7).fill = WHITE_FILL
        ws.cell(r, 7).font = BODY_FONT
        group_rows.append(r)
        r += 1

    # 소계
    g_rows_c = ",".join([f"C{x}" for x in group_rows])
    g_rows_d = ",".join([f"D{x}" for x in group_rows])
    g_rows_g = ",".join([f"G{x}" for x in group_rows])
    total(ws, r, 2, "  합  계"); ws.cell(r, 2).alignment = LEFT
    total(ws, r, 3, f"=SUM({g_rows_c})")
    total(ws, r, 4, f"=SUM({g_rows_d})")
    total(ws, r, 5, f"=C{r}+D{r}")
    ws.cell(r, 6).fill = SUM_FILL
    total(ws, r, 7, f"=SUM({g_rows_g})")
    sum_row = r; r += 2

    # 순수익 계산
    h2(ws, r, 2, "순 수익 계산 (결혼식 비용 차감)", 6); r += 1
    h3(ws, r, 2, "항  목"); h3(ws, r, 3, "금  액", 5); r += 1
    label(ws, r, 2, "예상 축의금 수입")
    inp(ws, r, 3, f"=G{sum_row}", NUM_FMT); r += 1
    label(ws, r, 2, "결혼식 총 비용 (💒 시트 연동)")
    inp(ws, r, 3, "='💒 결혼식_세부예산'!D100", NUM_FMT)
    ws.cell(r, 3).value = None  # 수동 입력
    r += 1
    label(ws, r, 2, "순 수익 (축의금 - 비용)")
    net_r = r
    total(ws, r, 3, f"=C{r-2}-C{r-1}"); r += 2

    # 시나리오
    h2(ws, r, 2, "시나리오별 예상", 6); r += 1
    h3(ws, r, 2, "시나리오")
    h3(ws, r, 3, "총 하객수")
    h3(ws, r, 4, "평균 축의금")
    h3(ws, r, 5, "예상 수입")
    h3(ws, r, 6, "비  고")
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    r += 1
    scenarios = [
        ("보수적", None, 80000, ""),
        ("기본",   None, 100000, ""),
        ("낙관적", None, 130000, ""),
    ]
    for s_name, guests, avg, note in scenarios:
        label(ws, r, 2, s_name)
        inp(ws, r, 3, guests)
        inp(ws, r, 4, avg)
        total(ws, r, 5, f"=C{r}*D{r}")
        ws.cell(r, 5).fill = WHITE_FILL; ws.cell(r, 5).font = BODY_FONT
        label(ws, r, 6, note); ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        r += 1

    r += 1
    ws.cell(r, 2, "  C·D열(하객수, 평균 축의금)을 입력하세요. 예상 축의금은 자동 계산됩니다.").font = Font(name="맑은 고딕", size=9, italic=True, color="808080")


# ══════════════════════════════════════════════════════════
# 3. 🏦 부동산_대출상세
# ══════════════════════════════════════════════════════════
def build_loan(wb):
    if "🏦 부동산_대출상세" in wb.sheetnames:
        del wb["🏦 부동산_대출상세"]
    ws = wb.create_sheet("🏦 부동산_대출상세")

    ws.column_dimensions["A"].width = 2
    set_col_width(ws, 2, 22)
    set_col_width(ws, 3, 18)
    set_col_width(ws, 4, 18)
    set_col_width(ws, 5, 14)
    set_col_width(ws, 6, 20)

    r = 2
    h1(ws, r, 2, "🏦  부동산 · 대출 상세 계획", 5); r += 1
    ws.row_dimensions[r-1].height = 30

    # ① 부동산 기본 정보
    h2(ws, r, 2, "① 부동산 기본 정보", 5); r += 1
    fields_basic = [
        ("아파트명 · 주소", None, "텍스트 입력"),
        ("매매가 (원)", None, ""),
        ("전용면적 (㎡)", None, ""),
        ("계약일", None, "YYYY-MM-DD"),
        ("잔금일", None, "YYYY-MM-DD"),
        ("입주 예정일", None, "YYYY-MM-DD"),
    ]
    for fname, fval, fnote in fields_basic:
        label(ws, r, 2, fname)
        c = inp(ws, r, 3, fval); c.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        label(ws, r, 5, fnote)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        r += 1

    r += 1
    # ② 자금 조달 계획
    h2(ws, r, 2, "② 자금 조달 계획", 5); r += 1
    h3(ws, r, 2, "항  목"); h3(ws, r, 3, "민  엽"); h3(ws, r, 4, "현  지"); h3(ws, r, 5, "합  계"); h3(ws, r, 6, "비  고"); r += 1

    fund_items = [
        "현금 (보유 자산)",
        "주식 매도 예정",
        "코인 매도 예정",
        "전세 보증금 회수",
        "부모님 지원",
        "기타",
    ]
    fund_rows = []
    for fi in fund_items:
        label(ws, r, 2, fi)
        inp(ws, r, 3); inp(ws, r, 4)
        total(ws, r, 5, f"=C{r}+D{r}")
        ws.cell(r, 5).fill = WHITE_FILL; ws.cell(r, 5).font = BODY_FONT
        label(ws, r, 6)
        fund_rows.append(r); r += 1

    fr_c = ",".join([f"C{x}" for x in fund_rows])
    fr_d = ",".join([f"D{x}" for x in fund_rows])
    total(ws, r, 2, "  자기 자금 합계"); ws.cell(r, 2).alignment = LEFT
    total(ws, r, 3, f"=SUM({fr_c})")
    total(ws, r, 4, f"=SUM({fr_d})")
    total(ws, r, 5, f"=C{r}+D{r}")
    self_fund_row = r; r += 2

    # ③ 대출 정보
    h2(ws, r, 2, "③ 대출 정보", 5); r += 1
    h3(ws, r, 2, "항  목"); h3(ws, r, 3, "대출 ①"); h3(ws, r, 4, "대출 ②"); h3(ws, r, 5, "합  계"); h3(ws, r, 6, "비  고"); r += 1

    loan_labels = [
        ("대출 종류", "", ""),
        ("대출 금액 (원)", None, None),
        ("금리 (%)", None, None),
        ("대출 기간 (년)", None, None),
        ("거치 기간 (년)", None, None),
        ("월 상환액 (원)", None, None),
        ("LTV (%)", None, None),
        ("DSR (%)", None, None),
        ("실행일", None, None),
        ("만기일", None, None),
    ]
    amt_rows = []
    monthly_rows = []
    for i, (ll, v1, v2) in enumerate(loan_labels):
        label(ws, r, 2, ll)
        c1 = inp(ws, r, 3, v1); c1.alignment = LEFT
        c2 = inp(ws, r, 4, v2); c2.alignment = LEFT
        if ll == "대출 금액 (원)":
            total(ws, r, 5, f"=C{r}+D{r}")
            ws.cell(r, 5).fill = WHITE_FILL; ws.cell(r, 5).font = BODY_FONT
            amt_rows.append(r)
        elif ll == "월 상환액 (원)":
            total(ws, r, 5, f"=C{r}+D{r}")
            ws.cell(r, 5).fill = WHITE_FILL; ws.cell(r, 5).font = BODY_FONT
            monthly_rows.append(r)
        else:
            ws.cell(r, 5).border = BORDER
        label(ws, r, 6)
        r += 1

    r += 1
    # ④ 취득세 · 부대비용
    h2(ws, r, 2, "④ 취득세 · 부대비용", 5); r += 1
    tax_items = [
        ("취득세 (매매가 × 세율)", None),
        ("농어촌특별세", None),
        ("지방교육세", None),
        ("법무사 수수료", None),
        ("중개 수수료", None),
        ("이사 비용", None),
        ("기타", None),
    ]
    tax_rows = []
    for ti, tv in tax_items:
        label(ws, r, 2, ti)
        c = inp(ws, r, 3, tv)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.cell(r, 5).border = BORDER; ws.cell(r, 6).border = BORDER
        tax_rows.append(r); r += 1

    tr_c = ",".join([f"C{x}" for x in tax_rows])
    total(ws, r, 2, "  취득세 합계"); ws.cell(r, 2).alignment = LEFT
    total(ws, r, 3, f"=SUM({tr_c})")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    tax_total_row = r; r += 2

    # ⑤ 인테리어 · 가전가구
    h2(ws, r, 2, "⑤ 인테리어 · 가전가구", 5); r += 1
    h3(ws, r, 2, "항  목"); h3(ws, r, 3, "예  산"); h3(ws, r, 4, "실제비용"); h3(ws, r, 5, "차  액"); h3(ws, r, 6, "비  고"); r += 1
    interior_items = [
        "철거 · 기본 공사", "도배 · 장판", "주방 공사", "욕실 공사",
        "전기 · 조명", "에어컨", "냉장고", "세탁기 · 건조기",
        "소파 · 침대 · 가구", "TV · 가전", "기타",
    ]
    int_rows = []
    for ii in interior_items:
        label(ws, r, 2, ii)
        inp(ws, r, 3); inp(ws, r, 4)
        inp(ws, r, 5, f"=D{r}-C{r}"); ws.cell(r, 5).font = BODY_FONT
        label(ws, r, 6)
        int_rows.append(r); r += 1

    ir_c = ",".join([f"C{x}" for x in int_rows])
    ir_d = ",".join([f"D{x}" for x in int_rows])
    total(ws, r, 2, "  인테리어 합계"); ws.cell(r, 2).alignment = LEFT
    total(ws, r, 3, f"=SUM({ir_c})")
    total(ws, r, 4, f"=SUM({ir_d})")
    total(ws, r, 5, f"=D{r}-C{r}")
    int_total_row = r; r += 2

    # ⑥ 전체 요약
    h1(ws, r, 2, "⑥ 총 자금 소요 요약", 5); r += 1
    summary = [
        ("매매가",              f"=C{fields_basic[1][0] if False else ''}"),  # 직접 링크 어려우니 수동
        ("자기 자금 합계",     f"=E{self_fund_row}"),
        ("대출 합계",          None),
        ("취득세 · 부대비용",  f"=C{tax_total_row}"),
        ("인테리어 · 가전가구",f"=C{int_total_row}"),
    ]
    for sn, sv in summary:
        label(ws, r, 2, sn)
        c = inp(ws, r, 3, sv if sv and "=C" not in sv else None)
        if sv and sv.startswith("="):
            c.value = sv
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.cell(r, 6).border = BORDER
        r += 1

    r += 1
    ws.cell(r, 2, "  노란색 셀에 금액을 직접 입력하세요. 대출 ①② 를 구분해 입력하면 합계가 자동 계산됩니다.").font = Font(name="맑은 고딕", size=9, italic=True, color="808080")


# ══════════════════════════════════════════════════════════
# 4. 📈 순자산_자동계산 (월별 추이 보완)
# ══════════════════════════════════════════════════════════
def build_net_worth(wb):
    if "📈 순자산_자동계산" in wb.sheetnames:
        del wb["📈 순자산_자동계산"]
    ws = wb.create_sheet("📈 순자산_자동계산")

    ws.column_dimensions["A"].width = 2
    set_col_width(ws, 2, 12)   # 월
    set_col_width(ws, 3, 14)   # 합계 수입
    set_col_width(ws, 4, 14)   # 총 지출
    set_col_width(ws, 5, 14)   # 저축액
    set_col_width(ws, 6, 16)   # 누적 저축
    set_col_width(ws, 7, 18)   # 순자산 (월말)
    set_col_width(ws, 8, 14)   # 저축률
    set_col_width(ws, 9, 22)   # 비고

    r = 2
    h1(ws, r, 2, "📈  월별 순자산 자동 계산", 8); r += 1
    ws.row_dimensions[r-1].height = 30

    # 기준 자산 입력
    h2(ws, r, 2, "기준 순자산 입력 (시작 시점)", 8); r += 1
    label(ws, r, 2, "시작 순자산 (원)")
    base_cell = inp(ws, r, 3, 450986632)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    label(ws, r, 5, "※ 대시보드 순자산 값을 입력")
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
    base_row = r; r += 2

    # 이벤트 차감 입력
    h2(ws, r, 2, "주요 이벤트 (해당 월에 차감)", 8); r += 1
    h3(ws, r, 2, "월"); h3(ws, r, 3, "이벤트명")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    h3(ws, r, 6, "차감액 (원)")
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    h3(ws, r, 8, "비  고")
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
    r += 1

    events = [
        ("2027-08", "결혼식 비용",     -50000000),
        ("2027-09", "부동산 잔금",     -200000000),
        ("2027-09", "취득세 등",       -50000000),
        ("2027-08", "인테리어/가전",   -40000000),
        ("2027-09", "신혼여행",        -10000000),
        ("2027-08", "전세보증금 회수", 200000000),
    ]
    event_rows = []
    for ev_month, ev_name, ev_amt in events:
        inp(ws, r, 2, ev_month, DATE_FMT); ws.cell(r, 2).alignment = LEFT; ws.cell(r, 2).border = BORDER
        label(ws, r, 3, ev_name)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        inp(ws, r, 6, ev_amt)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        label(ws, r, 8)
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
        event_rows.append(r); r += 1

    # 빈 이벤트 행 3개
    for _ in range(3):
        for col in [2, 3, 6, 8]:
            ws.cell(r, col).border = BORDER
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
        event_rows.append(r); r += 1

    event_section_end = r - 1
    r += 1

    # 월별 테이블
    h2(ws, r, 2, "월별 자산 추이 (수입 · 지출 입력)", 8); r += 1
    headers = ["월", "합계 수입", "총 지출", "저축액", "누적 저축", "순자산 (월말)", "저축률", "비  고"]
    for i, hdr in enumerate(headers):
        h3(ws, r, 2 + i, hdr)
    r += 1
    header_row = r - 1

    # 2026년 5월 ~ 2028년 12월 (32개월)
    import datetime
    months = []
    d = datetime.date(2026, 5, 1)
    while d <= datetime.date(2028, 12, 1):
        months.append(d)
        if d.month == 12:
            d = datetime.date(d.year + 1, 1, 1)
        else:
            d = datetime.date(d.year, d.month + 1, 1)

    data_start = r
    prev_net_row = None
    for idx, month in enumerate(months):
        # 월
        c_month = ws.cell(r, 2, month); c_month.number_format = DATE_FMT
        c_month.font = BODY_FONT; c_month.alignment = CENTER; c_month.border = BORDER

        inp(ws, r, 3)  # 합계 수입
        inp(ws, r, 4)  # 총 지출

        # 저축액 = 수입 - 지출
        c_sav = ws.cell(r, 5, f"=C{r}-D{r}")
        c_sav.number_format = NUM_FMT; c_sav.font = BODY_FONT
        c_sav.alignment = RIGHT; c_sav.border = BORDER

        # 누적 저축 = 이전 누적 + 이번 저축
        if idx == 0:
            c_cum = ws.cell(r, 6, f"=E{r}")
        else:
            c_cum = ws.cell(r, 6, f"=F{r-1}+E{r}")
        c_cum.number_format = NUM_FMT; c_cum.font = BODY_FONT
        c_cum.alignment = RIGHT; c_cum.border = BORDER

        # 이벤트 차감: 해당 월과 일치하는 이벤트 합산
        # SUMIF로 이벤트 시트 참조 (TEXT 비교)
        event_col6 = f"F{event_rows[0]}:F{event_rows[-1]}"
        event_col2 = f"B{event_rows[0]}:B{event_rows[-1]}"

        # 순자산 = 이전 순자산 + 저축 + 이벤트
        if idx == 0:
            c_net = ws.cell(r, 7,
                f"=C{base_row}+E{r}"
                f"+SUMPRODUCT(({event_col2}=TEXT(B{r},\"YYYY-MM\"))*{event_col6})")
        else:
            c_net = ws.cell(r, 7,
                f"=G{r-1}+E{r}"
                f"+SUMPRODUCT(({event_col2}=TEXT(B{r},\"YYYY-MM\"))*{event_col6})")
        c_net.number_format = NUM_FMT; c_net.font = BODY_FONT
        c_net.alignment = RIGHT; c_net.border = BORDER

        # 저축률
        c_pct = ws.cell(r, 8, f"=IF(C{r}=0,0,E{r}/C{r})")
        c_pct.number_format = PCT_FMT; c_pct.font = BODY_FONT
        c_pct.alignment = RIGHT; c_pct.border = BORDER

        label(ws, r, 9)  # 비고

        r += 1

    r += 1
    ws.cell(r, 2, "  C열(합계 수입), D열(총 지출)을 매월 입력하세요. 저축액·누적·순자산은 자동 계산됩니다.").font = Font(name="맑은 고딕", size=9, italic=True, color="808080")


# ══════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════
def main():
    print(f"파일 읽는 중: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    print("시트 추가 중: 💒 결혼식_세부예산")
    build_wedding(wb)

    print("시트 추가 중: 🎊 축의금_시뮬레이터")
    build_gift(wb)

    print("시트 추가 중: 🏦 부동산_대출상세")
    build_loan(wb)

    print("시트 추가 중: 📈 순자산_자동계산")
    build_net_worth(wb)

    wb.save(EXCEL_PATH)
    print(f"\n✅ 저장 완료: {EXCEL_PATH}")
    print("추가된 시트:")
    for name in wb.sheetnames:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
