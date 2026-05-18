"""
patch_v3c.py
자산현황 계약금 섹션 확장 (예식장계약금·취득세납부·기타 추가)
실행: python patch_v3c.py
"""
import sys, openpyxl
from copy import copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding="utf-8")
EXCEL_PATH = r"G:\내 드라이브\claude\자산관리\자산계획_v3.xlsx"

wb  = openpyxl.load_workbook(EXCEL_PATH)
ws  = wb["자산현황"]
wsd = wb["대시보드"]

# ── 현재 레이아웃 확인 ───────────────────────────────────
# R28: 계약금 납부 (아파트)  C28=130,000,000
# R29: 소계   =SUM(C28) / =SUM(E28)
# R30: 부채 섹션 시작 ...
# R36: 부채 소계  =SUM(C32:C35)
# R40: 순자산  =C16+C25+C29+C36
# R41: 가용자산  =C16+C25+C36

# ── Step 1: R29에 3행 삽입 → 새 R29,30,31 / 소계→R32 ──
print("R29에 3행 삽입 중...")
ws.insert_rows(29, 3)

# insert_rows 후 레이아웃:
# R28: 계약금(아파트)
# R29: [새 행] 예식장 계약금
# R30: [새 행] 취득세 납부
# R31: [새 행] 기타 납부
# R32: 소계 (자동이동)
# R35-38: 부채 데이터 (자동이동)
# R39: 부채 소계 (자동이동)
# R43: 순자산 (자동이동)
# R44: 가용자산 (자동이동)

# ── Step 2: 새 행 채우기 (R28 스타일 복사) ──────────────
def copy_cell(src, dst, value=None, formula=None):
    dst.font      = copy(src.font)
    dst.fill      = copy(src.fill)
    dst.border    = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    if formula is not None:
        dst.value = formula
    elif value is not None:
        dst.value = value

new_items = [
    (29, "예식장 계약금", "(예정)"),
    (30, "취득세 납부",   ""),
    (31, "기타 납부",     ""),
]

for (r, b_label, d_label) in new_items:
    ref = 28  # 스타일 복사 기준 행
    # B: 항목명
    copy_cell(ws.cell(ref, 2), ws.cell(r, 2), value=b_label)
    # C: 민엽 금액
    copy_cell(ws.cell(ref, 3), ws.cell(r, 3), value=0)
    ws.cell(r, 3).number_format = "#,##0"
    # D: 현지 항목명
    copy_cell(ws.cell(ref, 4) if ws.cell(ref, 4).value else ws.cell(ref, 2),
              ws.cell(r, 4), value=d_label)
    # E: 현지 금액
    copy_cell(ws.cell(ref, 5), ws.cell(r, 5), value=0)
    ws.cell(r, 5).number_format = "#,##0"
    # F: 합계
    copy_cell(ws.cell(ref, 6), ws.cell(r, 6), formula=f"=C{r}+E{r}")
    ws.cell(r, 6).number_format = "#,##0"

print(f"  새 행 R29,R30,R31 작성 완료")

# ── Step 3: 계약금 소계 수식 갱신 (R32) ─────────────────
# insert_rows 후 R32가 소계. SUM 범위를 C28:C31로 갱신
ws.cell(32, 3).value = "=SUM(C28:C31)"
ws.cell(32, 5).value = "=SUM(E28:E31)"
ws.cell(32, 6).value = "=C32+E32"
print("  R32 계약금 소계 수식 갱신")

# ── Step 4: 순자산·가용자산 수식 명시 업데이트 ───────────
# insert_rows가 in-sheet 참조를 자동 갱신하지만, 안전하게 명시 설정
# R43 = 순자산 (C29→C32, C36→C39)
ws.cell(43, 3).value = "=C16+C25+C32+C39"
ws.cell(43, 4).value = "=D16+D25+D32+D39"
ws.cell(43, 5).value = "=E16+E25+E32+E39"
ws.cell(43, 6).value = "=F16+F25+F32+F39"
# R44 = 가용자산 (C36→C39)
ws.cell(44, 3).value = "=C16+C25+C39"
ws.cell(44, 4).value = "=D16+D25+D39"
ws.cell(44, 5).value = "=E16+E25+E39"
ws.cell(44, 6).value = "=C44+E44"
print("  R43 순자산 / R44 가용자산 수식 갱신")

# ── Step 5: 대시보드 교차 참조 업데이트 ─────────────────
# R9: 계약금 소계 C29→C32
wsd.cell(9, 3).value = "=자산현황!C32"
wsd.cell(9, 4).value = "=자산현황!E32"
wsd.cell(9, 5).value = "=자산현황!F32"
# R10: 부채 소계 C36→C39
wsd.cell(10, 3).value = "=자산현황!C39"
wsd.cell(10, 4).value = "=자산현황!E39"
wsd.cell(10, 5).value = "=자산현황!F39"
# R11: 순자산 C40→C43
wsd.cell(11, 3).value = "=자산현황!C43"
wsd.cell(11, 4).value = "=자산현황!E43"
wsd.cell(11, 5).value = "=자산현황!F43"
# R23C6: 가용자산 F41→F44
wsd.cell(23, 6).value = "=자산현황!F44"
print("  대시보드 교차 참조 갱신")

wb.save(EXCEL_PATH)
print(f"\n완료! {EXCEL_PATH}")
print("  계약금 섹션: R28(아파트) + R29(예식장) + R30(취득세) + R31(기타) + R32(소계)")
print("  부채 섹션: R35-38 (데이터) / R39 (소계)")
print("  순자산: R43 / 가용자산: R44")
