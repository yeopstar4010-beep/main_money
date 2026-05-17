"""
민엽 & 현지 자산관리 엑셀 새로 생성
실행: python build_excel.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import datetime

OUT_PATH = r"G:\내 드라이브\claude\자산관리\자산계획_v2.xlsx"

# ── 컬러 팔레트 ───────────────────────────────────────────
C_NAVY   = "1F3864"
C_BLUE   = "2E75B6"
C_LBLUE  = "BDD7EE"
C_YELLOW = "FFF2CC"
C_GREEN  = "E2EFDA"
C_RED    = "FFE7E7"
C_GRAY   = "F2F2F2"
C_WHITE  = "FFFFFF"
C_DGRAY  = "595959"

# ── 스타일 헬퍼 ──────────────────────────────────────────
def fill(color): return PatternFill("solid", fgColor=color)
def font(color=C_NAVY, bold=False, size=10, italic=False):
    return Font(name="맑은 고딕", color=color, bold=bold, size=size, italic=italic)

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
BORDER_MED = Border(
    left=Side(style="medium", color="2E75B6"),
    right=Side(style="medium", color="2E75B6"),
    top=Side(style="medium", color="2E75B6"),
    bottom=Side(style="medium", color="2E75B6"),
)

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left",   vertical="center")
AL_R = Alignment(horizontal="right",  vertical="center")

NUM = '#,##0'
PCT = '0.0%'
DT  = 'YYYY-MM-DD'
YM  = 'YYYY-MM'

def w(ws, row, col, value=None, bg=None, fg=C_NAVY, bold=False, size=10,
      align=None, fmt=None, border=BORDER_THIN, cols=1, rows=1, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    if bg:    c.fill   = fill(bg)
    c.font   = font(fg, bold, size, italic)
    c.border = border
    if align: c.alignment = align
    if fmt:   c.number_format = fmt
    if cols > 1 or rows > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row+rows-1, end_column=col+cols-1)
    return c

def col_w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def row_h(ws, row, height):
    ws.row_dimensions[row].height = height

def section_title(ws, row, col, text, ncols=6):
    c = w(ws, row, col, f"  {text}", bg=C_NAVY, fg=C_WHITE, bold=True, size=11,
          align=AL_L, cols=ncols, border=Border())
    row_h(ws, row, 28)
    return c

def sub_title(ws, row, col, text, ncols=6):
    c = w(ws, row, col, f"  {text}", bg=C_BLUE, fg=C_WHITE, bold=True, size=10,
          align=AL_L, cols=ncols, border=Border())
    row_h(ws, row, 22)
    return c

def header(ws, row, col, text, ncols=1, nrows=1):
    return w(ws, row, col, text, bg=C_LBLUE, fg=C_NAVY, bold=True,
             align=AL_C, cols=ncols, rows=nrows)

def lbl(ws, row, col, text="", ncols=1):
    return w(ws, row, col, text, bg=C_GRAY, fg=C_DGRAY, align=AL_L, cols=ncols)

def inp(ws, row, col, value=None, fmt=NUM, ncols=1, color=C_WHITE):
    c = w(ws, row, col, value, bg=color, fg=C_NAVY, align=AL_R, fmt=fmt, cols=ncols)
    return c

def inp_l(ws, row, col, value=None, ncols=1):
    return w(ws, row, col, value, bg=C_WHITE, fg=C_NAVY, align=AL_L, cols=ncols)

def total_row(ws, row, col, value=None, fmt=NUM, ncols=1):
    return w(ws, row, col, value, bg=C_YELLOW, fg=C_NAVY, bold=True,
             align=AL_R, fmt=fmt, cols=ncols)

def note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="맑은 고딕", size=9, italic=True, color="808080")
    c.alignment = AL_L


# ═══════════════════════════════════════════════════════════
# 시트 1:  대시보드
# ═══════════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.create_sheet("대시보드")
    ws.sheet_view.showGridLines = False

    # 열 너비
    ws.column_dimensions["A"].width = 2
    for c, wd in [(2,18),(3,16),(4,16),(5,16),(6,4),(7,18),(8,16),(9,16)]:
        col_w(ws, c, wd)

    r = 2
    # ── 타이틀 ──
    w(ws, r, 2, "    민엽 & 현지  자산관리 대시보드", bg=C_NAVY, fg=C_WHITE,
      bold=True, size=14, align=AL_L, cols=8, border=Border())
    row_h(ws, r, 40); r += 1
    w(ws, r, 2, f"  기준일: {datetime.date.today().strftime('%Y년 %m월 %d일')}  |  각 시트에 입력하면 자동 업데이트",
      bg=C_LBLUE, fg=C_NAVY, align=AL_L, cols=8, border=Border())
    row_h(ws, r, 18); r += 2

    # ── 순자산 현황 ──
    sub_title(ws, r, 2, "현재 순자산", ncols=5); r += 1
    for col, txt in zip([2,3,4,5], ["항  목","민  엽","현  지","합  계"]):
        header(ws, r, col, txt)
    r += 1

    items = [
        ("현금·예금",       "='자산현황'!D5",  "='자산현황'!E5",  "='자산현황'!F5"),
        ("투자자산 (주식 등)","='자산현황'!D10", "='자산현황'!E10", "='자산현황'!F10"),
        ("계약금 납부",     "='자산현황'!D13", "='자산현황'!E13", "='자산현황'!F13"),
        ("부  채",          "='자산현황'!D17", "='자산현황'!E17", "='자산현황'!F17"),
    ]
    asset_rows = []
    for label_txt, f_m, f_h, f_t in items:
        lbl(ws, r, 2, label_txt)
        inp(ws, r, 3, f_m, NUM); inp(ws, r, 4, f_h, NUM); inp(ws, r, 5, f_t, NUM)
        asset_rows.append(r); r += 1

    lbl(ws, r, 2, "순  자  산")
    for col, cols_ref in [(3,"D"),(4,"E"),(5,"F")]:
        total_row(ws, r, col, f"='자산현황'!{cols_ref}19")
    net_worth_row = r; r += 2

    # ── D-Day ──
    sub_title(ws, r, 7, "D-Day", ncols=3); r += 1
    dday_items = [
        ("결혼식",    "2027-08-01", "='월별저축'!$D$3"),
        ("부동산 잔금", "2027-09-01", "='월별저축'!$D$3"),
    ]
    for ev, target, _ in dday_items:
        lbl(ws, r, 7, ev)
        inp(ws, r, 8, target, DT)
        total_row(ws, r, 9,
            f"=TEXT(DATEDIF(TODAY(),B{r},\"D\"),\"D-0일 남음\")",
            fmt="@")
        r += 1
    r += 1

    # ── 2027년 자금 계획 요약 ──
    sub_title(ws, r, 2, "2027년 자금 계획 요약", ncols=5); r += 1
    header(ws, r, 2, "항  목", ncols=2); header(ws, r, 4, "금  액"); header(ws, r, 5, "비  고"); r += 1

    plan_out = [
        ("잔금 (자기자금)",  "='부동산계획'!D7"),
        ("취득세·부대비용", "='부동산계획'!D21"),
        ("세입자 보증금",   "='부동산계획'!D25"),
        ("인테리어·가전",   "='부동산계획'!D34"),
        ("결혼식·신혼여행", "='결혼계획'!D36"),
    ]
    out_rows = []
    for pname, pf in plan_out:
        lbl(ws, r, 2, pname, ncols=2); inp(ws, r, 4, pf); inp_l(ws, r, 5)
        out_rows.append(r); r += 1
    out_ref = "+".join([f"D{x}" for x in out_rows])
    lbl(ws, r, 2, "지출 합계", ncols=2)
    total_row(ws, r, 4, f"={out_ref}"); r += 1

    plan_in = [
        ("현재 현금·투자 가용",  "='자산현황'!F20"),
        ("2027.9월까지 저축",    "='월별저축'!F35"),
        ("대출 예정",            "='부동산계획'!D12"),
    ]
    in_rows = []
    for pname, pf in plan_in:
        lbl(ws, r, 2, pname, ncols=2); inp(ws, r, 4, pf); inp_l(ws, r, 5)
        in_rows.append(r); r += 1
    in_ref = "+".join([f"D{x}" for x in in_rows])
    lbl(ws, r, 2, "가용 합계", ncols=2)
    total_row(ws, r, 4, f"={in_ref}")
    avail_row = r; r += 1

    lbl(ws, r, 2, "여유 / 부족", ncols=2)
    total_row(ws, r, 4, f"=D{avail_row}-D{avail_row-1}")
    ws.cell(r, 4).fill = fill(C_GREEN); r += 2

    note(ws, r, 2, "  ※ 노란색 셀은 각 시트에서 자동으로 불러옵니다. 각 시트에 직접 입력해주세요.")

    return ws


# ═══════════════════════════════════════════════════════════
# 시트 2:  자산현황
# ═══════════════════════════════════════════════════════════
def build_assets(wb):
    ws = wb.create_sheet("자산현황")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    for c, wd in [(2,22),(3,4),(4,18),(5,18),(6,18),(7,20)]:
        col_w(ws, c, wd)

    r = 2
    w(ws, r, 2, "    자산 현황", bg=C_NAVY, fg=C_WHITE, bold=True, size=13,
      align=AL_L, cols=6, border=Border())
    row_h(ws, r, 36); r += 1
    w(ws, r, 2, f"  기준: 매월 말일 업데이트", bg=C_LBLUE, fg=C_NAVY,
      align=AL_L, cols=6, border=Border())
    row_h(ws, r, 16); r += 2

    header(ws, r, 2, "항  목"); w(ws, r, 3, ""); header(ws, r, 4, "민  엽")
    header(ws, r, 5, "현  지"); header(ws, r, 6, "합  계")
    w(ws, r, 7, "비  고", bg=C_LBLUE, fg=C_NAVY, bold=True, align=AL_C)
    r += 1

    def asset_section(title, items_m, items_h, sum_row_label):
        nonlocal r
        sub_title(ws, r, 2, title, ncols=6); r += 1
        item_rows = []
        for im, ih in zip(items_m, items_h):
            lbl(ws, r, 2, im[0]); w(ws, r, 3, "")
            inp(ws, r, 4, im[1])
            inp(ws, r, 5, ih[1])
            c = ws.cell(r, 6, f"=D{r}+E{r}")
            c.fill = fill(C_WHITE); c.font = font(); c.border = BORDER_THIN
            c.number_format = NUM; c.alignment = AL_R
            inp_l(ws, r, 7, im[2] if len(im)>2 else "")
            item_rows.append(r); r += 1
        sum_ref_d = "+".join([f"D{x}" for x in item_rows])
        sum_ref_e = "+".join([f"E{x}" for x in item_rows])
        lbl(ws, r, 2, sum_row_label); w(ws, r, 3, "")
        total_row(ws, r, 4, f"=SUM({sum_ref_d})")
        total_row(ws, r, 5, f"=SUM({sum_ref_e})")
        total_row(ws, r, 6, f"=D{r}+E{r}")
        w(ws, r, 7, ""); row_h(ws, r, 20)
        result_row = r; r += 1
        return result_row

    # 현금·예금
    cash_m = [
        ("토스뱅크", 1794954), ("하나은행", 72561), ("기업은행", 109094),
        ("카카오뱅크", 150012), ("우리은행", 50000), ("OK토스플러스", 300000),
        ("지역화폐", 36779), ("배터리카드", 400000), ("기타", 3600000),
    ]
    cash_h = [
        ("OK저축은행", 20722500), ("주택청약", 7000000), ("농협", 766576),
        ("적금 ①", 0), ("적금 ②", 0), ("기타 예금", 0),
        ("", 0), ("", 0), ("", 0),
    ]
    cash_total = asset_section("현금 · 예금", cash_m, cash_h, "  현금·예금 소계")

    # 투자자산
    inv_m = [
        ("키움증권", 87576544), ("토스주식", 99471104),
        ("자사주", 9381394), ("ISA (민엽)", 9381394),
        ("IRP·연금저축", 21144836), ("코인", 31263027),
    ]
    inv_h = [
        ("한국투자증권", 76114704), ("토스증권", 41075120),
        ("대신증권", 2035000), ("ISA (현지)", 5297783),
        ("연금저축 (현지)", 2562952), ("업비트", 680298),
    ]
    inv_total = asset_section("투자 자산 (주식 · 코인 등)", inv_m, inv_h, "  투자자산 소계")

    # 계약금 납부
    cont_m = [("계약금 납부 (아파트)", 130000000, "2025.04 납부, 잔금 전까지 고정")]
    cont_h = [("", 0, "")]
    contract_total = asset_section("계약금 납부 (유동성 없음)", cont_m, cont_h, "  계약금 소계")

    # 부채
    debt_m = [("어머니 차용금", -106000000, "계약금 조달 시 차용")]
    debt_h = [("", 0, "")]
    debt_total = asset_section("부  채 (마이너스)", debt_m, debt_h, "  부채 소계")

    # 순자산 합계
    r += 1
    sub_title(ws, r, 2, "순  자  산  합  계", ncols=6); r += 1
    lbl(ws, r, 2, "순자산 = 현금+투자+계약금 - 부채"); w(ws, r, 3, "")
    for col, ref in [(4, "D"),(5, "E"),(6, "F")]:
        total_row(ws, r, col,
            f"={ref}{cash_total}+{ref}{inv_total}+{ref}{contract_total}+{ref}{debt_total}")
        ws.cell(r, col).font = font(C_NAVY, bold=True, size=11)
    row_h(ws, r, 24); r += 2

    # 가용 가능 자산 (계약금 제외)
    sub_title(ws, r, 2, "실제 가용 자산 (계약금 제외)", ncols=6); r += 1
    note(ws, r, 2, "  ※ 부동산 잔금·결혼 비용 마련을 위해 실제 쓸 수 있는 금액")
    r += 1
    lbl(ws, r, 2, "가용 자산 = 현금+투자 - 부채"); w(ws, r, 3, "")
    for col, ref in [(4, "D"),(5, "E"),(6, "F")]:
        total_row(ws, r, col,
            f"={ref}{cash_total}+{ref}{inv_total}+{ref}{debt_total}")
    row_h(ws, r, 24); r += 2

    note(ws, r, 2, "  ※ D·E열(금액)만 수정하세요. F열 합계와 소계는 자동 계산됩니다.")

    return ws


# ═══════════════════════════════════════════════════════════
# 시트 3:  월별저축
# ═══════════════════════════════════════════════════════════
def build_monthly(wb):
    ws = wb.create_sheet("월별저축")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    for c, wd in [(2,10),(3,14),(4,14),(5,14),(6,14),(7,16),(8,12),(9,22)]:
        col_w(ws, c, wd)

    r = 2
    w(ws, r, 2, "    월별 저축 추이", bg=C_NAVY, fg=C_WHITE, bold=True, size=13,
      align=AL_L, cols=8, border=Border())
    row_h(ws, r, 36); r += 1

    # 기준 순자산 입력
    sub_title(ws, r, 2, "시작 순자산 (2026년 5월 기준)", ncols=8); r += 1
    lbl(ws, r, 2, "시작 순자산")
    base_cell_row = r
    inp(ws, r, 3, 450986632, ncols=2)
    inp_l(ws, r, 5, "  ←  자산현황 시트 순자산 값", ncols=4)
    r += 2

    # 이벤트 정의
    sub_title(ws, r, 2, "주요 이벤트 (자동 반영)", ncols=8); r += 1
    header(ws, r, 2, "연-월"); header(ws, r, 3, "이벤트명", ncols=3)
    header(ws, r, 6, "금액 (원)"); header(ws, r, 7, "구분"); header(ws, r, 8, "비  고")
    r += 1

    events = [
        ("2027-08", "어머니 차용금 상환",   -106000000, "지출"),
        ("2027-08", "결혼식 비용",           -50000000,  "지출"),
        ("2027-08", "인테리어·가전가구",     -40000000,  "지출"),
        ("2027-09", "부동산 잔금 (자기자금)",-220000000, "지출"),
        ("2027-09", "취득세·부대비용",       -50000000,  "지출"),
        ("2027-09", "세입자 보증금 반환",    -300000000, "지출"),
        ("2027-09", "신혼여행",              -10000000,  "지출"),
    ]
    event_rows = []
    for ev_ym, ev_name, ev_amt, ev_type in events:
        inp_l(ws, r, 2, ev_ym)
        inp_l(ws, r, 3, ev_name, ncols=3)
        inp(ws, r, 6, ev_amt, color=C_RED if ev_amt < 0 else C_GREEN)
        inp_l(ws, r, 7, ev_type)
        inp_l(ws, r, 8)
        event_rows.append(r); r += 1
    # 빈 이벤트 행 3개
    for _ in range(3):
        inp_l(ws, r, 2); inp_l(ws, r, 3, ncols=3); inp(ws, r, 6); inp_l(ws, r, 7); inp_l(ws, r, 8)
        event_rows.append(r); r += 1
    event_end = r - 1; r += 1

    # 월별 테이블
    sub_title(ws, r, 2, "월별 기록 (수입·지출 입력)", ncols=8); r += 1
    for col, txt in zip([2,3,4,5,6,7,8,9],
                        ["월","민엽 수입","현지 수입","합계 수입","총 지출","저축액","저축률","비  고"]):
        header(ws, r, col, txt)
    r += 1
    data_start = r

    months = []
    d = datetime.date(2026, 5, 1)
    while d <= datetime.date(2028, 12, 1):
        months.append(d)
        d = datetime.date(d.year + (d.month == 12), (d.month % 12) + 1, 1)

    preset = {
        "2026-05": (6000000, 3500000, 4000000),
        "2026-06": (6000000, 3500000, 4000000),
        "2026-07": (6000000, 3500000, 4000000),
        "2026-08": (6000000, 5500000, 5500000),
        "2026-09": (6000000, 3500000, 5500000),
        "2026-10": (6000000, 3500000, 5500000),
        "2026-11": (6000000, 3500000, 5500000),
        "2026-12": (6000000, 3500000, 5500000),
        "2027-01": (16000000, 3500000, 5500000),
        "2027-02": (6000000, 3500000, 5500000),
        "2027-03": (6000000, 3500000, 5500000),
        "2027-04": (6000000, 3500000, 5500000),
        "2027-05": (6000000, 3500000, 5500000),
        "2027-06": (6000000, 3500000, 5500000),
        "2027-07": (6000000, 3500000, 5500000),
        "2027-08": (6000000, 5500000, 5500000),
        "2027-09": (6000000, 3500000, 5500000),
    }

    prev_net = None
    for idx, month in enumerate(months):
        ym = month.strftime("%Y-%m")
        p = preset.get(ym, (None, None, None))

        c_m = ws.cell(r, 2, month); c_m.number_format = YM
        c_m.font = font(); c_m.alignment = AL_C; c_m.border = BORDER_THIN

        inp(ws, r, 3, p[0]); inp(ws, r, 4, p[1])

        # 합계 수입
        c_tot = ws.cell(r, 5, f"=C{r}+D{r}")
        c_tot.fill = fill(C_WHITE); c_tot.font = font(); c_tot.border = BORDER_THIN
        c_tot.number_format = NUM; c_tot.alignment = AL_R

        inp(ws, r, 6, p[2])

        # 저축
        c_sav = ws.cell(r, 7, f"=E{r}-F{r}")
        c_sav.fill = fill(C_WHITE); c_sav.font = font(); c_sav.border = BORDER_THIN
        c_sav.number_format = NUM; c_sav.alignment = AL_R

        # 저축률
        c_pct = ws.cell(r, 8, f"=IF(E{r}=0,0,G{r}/E{r})")
        c_pct.fill = fill(C_WHITE); c_pct.font = font(); c_pct.border = BORDER_THIN
        c_pct.number_format = PCT; c_pct.alignment = AL_R

        inp_l(ws, r, 9)
        r += 1

    data_end = r - 1

    # 합계 행
    r += 1
    sub_title(ws, r, 2, "누계 / 평균", ncols=8); r += 1
    lbl(ws, r, 2, "2026.05 ~ 2027.09 저축 합계")
    total_row(ws, r, 7, f"=SUMPRODUCT((B{data_start}:B{data_end}>=DATE(2026,5,1))*(B{data_start}:B{data_end}<=DATE(2027,9,1))*G{data_start}:G{data_end})")
    r += 1
    lbl(ws, r, 2, "평균 월 저축")
    total_row(ws, r, 7, f"=AVERAGE(G{data_start}:G{data_end})")
    r += 2

    note(ws, r, 2, "  ※ C·D열(수입), F열(지출)을 매월 입력하세요. G·H열은 자동 계산됩니다.")
    note(ws, r+1, 2, "  ※ 이벤트는 위 '주요 이벤트' 표에서 수정하세요.")

    return ws, base_cell_row, event_rows, data_start, data_end


# ═══════════════════════════════════════════════════════════
# 시트 4:  부동산계획
# ═══════════════════════════════════════════════════════════
def build_real_estate(wb):
    ws = wb.create_sheet("부동산계획")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    for c, wd in [(2,24),(3,4),(4,18),(5,18),(6,20)]:
        col_w(ws, c, wd)

    r = 2
    w(ws, r, 2, "    부동산 계획", bg=C_NAVY, fg=C_WHITE, bold=True, size=13,
      align=AL_L, cols=5, border=Border())
    row_h(ws, r, 36); r += 1
    w(ws, r, 2, "  노란색 셀에 값을 입력하세요.", bg=C_YELLOW, fg=C_NAVY,
      align=AL_L, cols=5, border=Border())
    row_h(ws, r, 16); r += 2

    # ① 기본 정보
    sub_title(ws, r, 2, "① 부동산 기본 정보", ncols=5); r += 1
    basics = [
        ("아파트명·단지", None, "텍스트 입력"),
        ("주소", None, ""),
        ("총 매매가 (원)", 1210000000, ""),
        ("계약일", "2025-04-17", ""),
        ("잔금 예정일", "2027-09-01", ""),
        ("입주 예정일", "2027-09-01", ""),
    ]
    for bname, bval, bnote in basics:
        lbl(ws, r, 2, bname); w(ws, r, 3, "")
        if isinstance(bval, int):
            inp(ws, r, 4, bval, ncols=2)
        else:
            inp_l(ws, r, 4, bval, ncols=2)
        inp_l(ws, r, 6, bnote)
        r += 1
    r += 1

    # ② 자금 조달 계획
    sub_title(ws, r, 2, "② 자금 조달 계획", ncols=5); r += 1
    header(ws, r, 2, "항  목"); w(ws, r, 3, "")
    header(ws, r, 4, "금  액"); header(ws, r, 5, "비  고", ncols=2); r += 1
    fund_items = [
        ("계약금 (이미 납부)",     130000000, "현금 1천만 + 계약금 1억 2천"),
        ("잔금 자기자금",          220000000, ""),
        ("대출 예정금액",          560000000, "LTV·DSR 확인 후 수정"),
        ("기타 (부모님 지원 등)",  0,         ""),
    ]
    fund_rows = []
    for fname, fval, fnote in fund_items:
        lbl(ws, r, 2, fname); w(ws, r, 3, "")
        inp(ws, r, 4, fval); inp_l(ws, r, 5, fnote, ncols=2)
        fund_rows.append(r); r += 1
    fr = "+".join([f"D{x}" for x in fund_rows])
    lbl(ws, r, 2, "합  계"); w(ws, r, 3, "")
    total_row(ws, r, 4, f"=SUM({fr})")
    inp_l(ws, r, 5, "≒ 총 매매가와 일치 확인", ncols=2); r += 2

    # ③ 대출 상세
    sub_title(ws, r, 2, "③ 대출 상세", ncols=5); r += 1
    loan_items = [
        ("대출 종류", None, "예: 주택담보대출, 디딤돌 등"),
        ("대출 금액 (원)", None, ""),
        ("연 금리 (%)", None, ""),
        ("대출 기간 (년)", None, ""),
        ("거치 기간 (년)", None, "0이면 즉시 원리금 상환"),
        ("월 상환액 (원)", None, "은행 계산기로 확인"),
        ("실행 예정일", None, ""),
        ("만기일", None, ""),
    ]
    for lname, lval, lnote in loan_items:
        lbl(ws, r, 2, lname); w(ws, r, 3, "")
        inp_l(ws, r, 4, lval); inp_l(ws, r, 5, lnote, ncols=2)
        r += 1
    r += 1

    # ④ 취득세·부대비용
    sub_title(ws, r, 2, "④ 취득세 · 부대비용", ncols=5); r += 1
    header(ws, r, 2, "항  목"); w(ws, r, 3, "")
    header(ws, r, 4, "예산"); header(ws, r, 5, "실제"); header(ws, r, 6, "비  고"); r += 1
    tax_items = [
        ("취득세",        None),
        ("농어촌특별세",  None),
        ("지방교육세",    None),
        ("법무사 수수료", None),
        ("부동산 중개수수료", None),
        ("이사 비용",     None),
        ("기타",          None),
    ]
    tax_rows = []
    for tname, tval in tax_items:
        lbl(ws, r, 2, tname); w(ws, r, 3, "")
        inp(ws, r, 4, tval); inp(ws, r, 5); inp_l(ws, r, 6)
        tax_rows.append(r); r += 1
    tr = "+".join([f"D{x}" for x in tax_rows])
    lbl(ws, r, 2, "취득세 합계"); w(ws, r, 3, "")
    total_row(ws, r, 4, f"=SUM({tr})")
    r += 2

    # ⑤ 세입자 보증금
    sub_title(ws, r, 2, "⑤ 세입자 보증금 (반환해야 할 금액)", ncols=5); r += 1
    lbl(ws, r, 2, "현 세입자 보증금"); w(ws, r, 3, "")
    inp(ws, r, 4, 300000000)
    inp_l(ws, r, 5, "입주 시 세입자에게 반환", ncols=2); r += 1
    lbl(ws, r, 2, "반환 예정일"); w(ws, r, 3, "")
    inp_l(ws, r, 4, "2027-09-01"); inp_l(ws, r, 5, "", ncols=2); r += 2

    # ⑥ 인테리어·가전가구
    sub_title(ws, r, 2, "⑥ 인테리어 · 가전가구", ncols=5); r += 1
    header(ws, r, 2, "항  목"); w(ws, r, 3, "")
    header(ws, r, 4, "예산"); header(ws, r, 5, "실제"); header(ws, r, 6, "비  고"); r += 1
    interior_items = [
        "철거·기본 공사", "도배·장판", "주방 공사", "욕실 공사",
        "전기·조명", "에어컨", "냉장고", "세탁기·건조기",
        "소파·침대·가구", "TV·가전", "기타",
    ]
    int_rows = []
    for iname in interior_items:
        lbl(ws, r, 2, iname); w(ws, r, 3, "")
        inp(ws, r, 4); inp(ws, r, 5); inp_l(ws, r, 6)
        int_rows.append(r); r += 1
    ir = "+".join([f"D{x}" for x in int_rows])
    lbl(ws, r, 2, "인테리어 합계"); w(ws, r, 3, "")
    total_row(ws, r, 4, f"=SUM({ir})"); r += 2

    note(ws, r, 2, "  ※ D열(예산)을 입력하세요. 실제 비용은 계약/구매 후 E열에 업데이트.")

    return ws


# ═══════════════════════════════════════════════════════════
# 시트 5:  결혼계획
# ═══════════════════════════════════════════════════════════
def build_wedding(wb):
    ws = wb.create_sheet("결혼계획")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    for c, wd in [(2,22),(3,4),(4,16),(5,16),(6,14),(7,18)]:
        col_w(ws, c, wd)

    r = 2
    w(ws, r, 2, "    결혼 계획", bg=C_NAVY, fg=C_WHITE, bold=True, size=13,
      align=AL_L, cols=6, border=Border())
    row_h(ws, r, 36); r += 1
    w(ws, r, 2, "  결혼식 예정: 2027년 8월  |  신혼여행 예정: 2027년 9월", bg=C_LBLUE,
      fg=C_NAVY, align=AL_L, cols=6, border=Border())
    row_h(ws, r, 16); r += 2

    # ① 결혼식 세부예산
    sub_title(ws, r, 2, "① 결혼식 세부 예산", ncols=6); r += 1
    header(ws, r, 2, "항  목"); w(ws, r, 3, "")
    header(ws, r, 4, "예산 (원)"); header(ws, r, 5, "실제 (원)")
    header(ws, r, 6, "차  액"); header(ws, r, 7, "비  고"); r += 1

    wedding_sections = [
        ("스드메", ["스튜디오 (사진 촬영)", "드레스 (웨딩드레스)", "메이크업 (헤어 포함)", "신랑 예복", "부케·부토니에"]),
        ("예식장", ["예식장 대관료", "뷔페·식사", "웨딩 케이크", "사회자", "영상 촬영", "청첩장·답례품"]),
        ("예물·예단", ["예물 (반지 등)", "예단 비용"]),
        ("기타", ["기타 ①", "기타 ②"]),
    ]

    all_budget_rows = []
    for sec_name, items in wedding_sections:
        sec_rows = []
        for item in items:
            lbl(ws, r, 2, f"  {item}"); w(ws, r, 3, "")
            inp(ws, r, 4); inp(ws, r, 5)
            c_diff = ws.cell(r, 6, f"=E{r}-D{r}")
            c_diff.fill = fill(C_WHITE); c_diff.font = font()
            c_diff.border = BORDER_THIN; c_diff.number_format = NUM; c_diff.alignment = AL_R
            inp_l(ws, r, 7)
            sec_rows.append(r); all_budget_rows.append(r); r += 1
        sec_ref_d = "+".join([f"D{x}" for x in sec_rows])
        sec_ref_e = "+".join([f"E{x}" for x in sec_rows])
        lbl(ws, r, 2, f"  [{sec_name}] 소계"); w(ws, r, 3, "")
        total_row(ws, r, 4, f"=SUM({sec_ref_d})")
        total_row(ws, r, 5, f"=SUM({sec_ref_e})")
        total_row(ws, r, 6, f"=E{r}-D{r}")
        w(ws, r, 7, ""); r += 1

    all_d = "+".join([f"D{x}" for x in all_budget_rows])
    all_e = "+".join([f"E{x}" for x in all_budget_rows])
    r += 1
    lbl(ws, r, 2, "결혼식 합계"); w(ws, r, 3, "")
    total_row(ws, r, 4, f"=SUM({all_d})")
    total_row(ws, r, 5, f"=SUM({all_e})")
    total_row(ws, r, 6, f"=E{r}-D{r}")
    wedding_total_row = r; r += 2

    # ② 신혼여행
    sub_title(ws, r, 2, "② 신혼여행", ncols=6); r += 1
    header(ws, r, 2, "항  목"); w(ws, r, 3, "")
    header(ws, r, 4, "예산"); header(ws, r, 5, "실제")
    header(ws, r, 6, "차  액"); header(ws, r, 7, "비  고"); r += 1
    trip_items = ["항공권 (2인)", "숙박", "투어·액티비티", "식비·쇼핑", "기타"]
    trip_rows = []
    for tname in trip_items:
        lbl(ws, r, 2, f"  {tname}"); w(ws, r, 3, "")
        inp(ws, r, 4); inp(ws, r, 5)
        c = ws.cell(r, 6, f"=E{r}-D{r}")
        c.fill = fill(C_WHITE); c.font = font(); c.border = BORDER_THIN
        c.number_format = NUM; c.alignment = AL_R
        inp_l(ws, r, 7)
        trip_rows.append(r); r += 1
    tr_d = "+".join([f"D{x}" for x in trip_rows])
    tr_e = "+".join([f"E{x}" for x in trip_rows])
    lbl(ws, r, 2, "신혼여행 합계"); w(ws, r, 3, "")
    total_row(ws, r, 4, f"=SUM({tr_d})")
    total_row(ws, r, 5, f"=SUM({tr_e})")
    total_row(ws, r, 6, f"=E{r}-D{r}")
    trip_total_row = r; r += 2

    # ③ 결혼 총합
    sub_title(ws, r, 2, "③ 결혼 비용 총합", ncols=6); r += 1
    lbl(ws, r, 2, "결혼식 + 신혼여행 합계"); w(ws, r, 3, "")
    total_row(ws, r, 4, f"=D{wedding_total_row}+D{trip_total_row}")
    total_row(ws, r, 5, f"=E{wedding_total_row}+E{trip_total_row}")
    total_row(ws, r, 6, f"=E{r}-D{r}")
    ws.cell(r, 4).font = font(C_NAVY, bold=True, size=11)
    row_h(ws, r, 24); grand_total_row = r; r += 2

    # ④ 축의금 시뮬레이터
    sub_title(ws, r, 2, "④ 축의금 시뮬레이터", ncols=6); r += 1
    header(ws, r, 2, "하객 구분"); w(ws, r, 3, "")
    header(ws, r, 4, "민엽 인원"); header(ws, r, 5, "현지 인원")
    header(ws, r, 6, "평균 축의금"); header(ws, r, 7, "예상 수입"); r += 1
    gift_groups = [
        ("직장 동료", None, None, 70000),
        ("친구·지인", None, None, 100000),
        ("친척·가족", None, None, 200000),
        ("기타",       None, None, 100000),
    ]
    gift_rows = []
    for gname, gm, gh, gavg in gift_groups:
        lbl(ws, r, 2, gname); w(ws, r, 3, "")
        inp(ws, r, 4, gm); inp(ws, r, 5, gh); inp(ws, r, 6, gavg)
        c = ws.cell(r, 7, f"=(D{r}+E{r})*F{r}")
        c.fill = fill(C_WHITE); c.font = font(); c.border = BORDER_THIN
        c.number_format = NUM; c.alignment = AL_R
        gift_rows.append(r); r += 1
    gr_g = "+".join([f"G{x}" for x in gift_rows])
    lbl(ws, r, 2, "축의금 합계"); w(ws, r, 3, "")
    for col in [4,5,6]: ws.cell(r, col).border = BORDER_THIN
    total_row(ws, r, 7, f"=SUM({gr_g})")
    gift_total_row = r; r += 1

    lbl(ws, r, 2, "결혼 순비용 (총비용 - 축의금)"); w(ws, r, 3, "")
    for col in [4,5,6]: ws.cell(r, col).border = BORDER_THIN
    total_row(ws, r, 7, f"=D{grand_total_row}-G{gift_total_row}")
    r += 2

    note(ws, r, 2, "  ※ D열(예산), 하객 인원을 입력하세요. 실제 비용은 계약 후 E열에 업데이트.")

    return ws


# ═══════════════════════════════════════════════════════════
# 메인
# ═══════════════════════════════════════════════════════════
def main():
    import sys
    sys.stdout.reconfigure(encoding="utf-8")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    print("시트 생성 중...")
    build_dashboard(wb)
    print("  ✓ 대시보드")
    build_assets(wb)
    print("  ✓ 자산현황")
    build_monthly(wb)
    print("  ✓ 월별저축")
    build_real_estate(wb)
    print("  ✓ 부동산계획")
    build_wedding(wb)
    print("  ✓ 결혼계획")

    wb.save(OUT_PATH)
    print(f"\n✅ 저장 완료: {OUT_PATH}")


if __name__ == "__main__":
    main()
